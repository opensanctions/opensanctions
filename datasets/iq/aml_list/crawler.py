import re
from openpyxl import load_workbook
from rigour.mime.types import XLSX
from typing import Optional

from zavod import Context, helpers as h
from zavod.shed.zyte_api import fetch_html, fetch_resource

# These sheets do not contain actual data but serve as reference sheets
LOC_IGNORE_LIST = [
    "الافراد",  # Individuals
    "القوائم المحلية ",  # Local lists
]

YEAR_PATTERN = re.compile(r"\b\d{4}\b")
# To mediate in the sale and purchase of foreign currencies
ENTITY_NAME_REASON = re.compile(r"\s*للتوسط ببيع وشراء العملات الاجنبية$")


def clean_entity_name(entity_name: str) -> Optional[str]:
    if entity_name is None:
        return None
    return ENTITY_NAME_REASON.sub("", entity_name).strip()


def extract_sector(entity_name: str) -> Optional[str]:
    if entity_name is None:
        return None
    match = ENTITY_NAME_REASON.search(entity_name)
    return match.group().strip() if match else None


def extract_listing_date(decision_number: str) -> Optional[str]:
    """Extracts the listing year from the decision number."""
    if decision_number is None:
        return None
    match = YEAR_PATTERN.search(decision_number)
    return match.group(0) if match else None


def crawl_row(row: dict, context: Context):
    raw_entity_name = row.pop("entity_name", None)
    decision_number = row.pop("decision_no")
    entity_name = clean_entity_name(raw_entity_name)
    listing_date = extract_listing_date(decision_number)

    if entity_name:
        entity = context.make("LegalEntity")
        entity.id = context.make_id(entity_name, decision_number)
        entity.add("name", entity_name, lang="ara")
        entity.add("sector", extract_sector(raw_entity_name), lang="ara")
    else:
        name = row.pop("name", row.pop("person_name"))
        birth_date = row.pop("dob")
        entity = context.make("Person")
        entity.id = context.make_id(name, birth_date)
        entity.add("nationality", row.pop("nationality", None), lang="ara")
        h.apply_date(entity, "birthDate", birth_date)
        h.apply_name(
            entity,
            full=name,
            matronymic=row.pop("matronymic"),
            lang="ara",
        )

    entity.add("topics", "sanction")

    sanction = h.make_sanction(context, entity)
    sanction.add("recordId", decision_number)
    h.apply_date(sanction, "listingDate", listing_date)

    context.emit(entity)
    context.emit(sanction)

    context.audit_data(row, ["id"])


def process_xlsx(
    context,
    url: str,
    filename: str,
    title: str,
    ignore_sheets: list = [],
):
    excel_link_xpath = (
        '//article[contains(@id, "post-")]//a[contains(@href, "xlsx")]/@href'
    )
    doc = fetch_html(context, url, excel_link_xpath, cache_days=1, geolocation="IQ")
    link = doc.xpath(excel_link_xpath)

    assert len(link) == 1, link
    file_url = link[0]
    assert file_url.endswith(".xlsx"), file_url
    assert title in file_url, file_url

    _, _, _, path = fetch_resource(context, filename, file_url, XLSX, geolocation="IQ")
    context.export_resource(path, XLSX)

    wb = load_workbook(path, read_only=True)
    processed_sheets = set()
    for sheet in wb.sheetnames:
        if sheet in ignore_sheets:
            continue
        for row in h.parse_xlsx_sheet(
            context, wb[sheet], skiprows=3, header_lookup="columns"
        ):
            crawl_row(row, context)
            processed_sheets.add(sheet)

    assert set(wb.sheetnames) == processed_sheets | set(ignore_sheets)


def crawl(context: Context):
    process_xlsx(
        context,
        "https://aml.iq/?page_id=2169",
        "international.xlsx",
        "القائمة-الدولية",  # International list
    )
    process_xlsx(
        context,
        "https://aml.iq/?page_id=2171",
        "local.xlsx",
        "القوائم-المحلية",  # Local lists
        LOC_IGNORE_LIST,
    )
