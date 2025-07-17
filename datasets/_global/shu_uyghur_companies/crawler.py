import openpyxl

from zavod import Context, helpers as h

OPERATING_URL = "https://data.opensanctions.org/contrib/shu_uyghur_companies/Companies%20Operating%20in%20XUAR-web.archive.org-20240331100003.xlsx"
LABOUR_TRANSFERS_URL = "https://data.opensanctions.org/contrib/shu_uyghur_companies/2023-07-03%20Companies%20Named%20in%20Reports%20v2-web.archive.org-20250102144725.xlsx"

LABOUR_SHEETS = [
    "1. Labor Transfers in XUAR",
    "2. Labor Transfers Out of XUAR",
    "3. XPCC",
    "4. Priority Sector Apparel",
    "5. Priority Sector Tomato",
    "6. Priority Sector Polysilicon",
]
APP_LABOUR_SHEETS = LABOUR_SHEETS + ["Information"]

OPERATING_SHEETS = {
    "Info ",
    "1. Companies Operating in XUAR",
    "2. Export Relevant Categories",
}


def apply_addresses(context, entity, addr, addr_en, city, city_en):
    """Create and apply addresses to an entity."""
    if addr:
        address_ent = h.make_address(context, full=addr, city=city, lang="zhu")
        h.copy_address(entity, address_ent)
    if addr_en:
        address_en_ent = h.make_address(context, full=addr_en, city=city_en, lang="eng")
        h.copy_address(entity, address_en_ent)


def crawl_labour_transfers(context: Context, labour_transfers_url):
    path = context.fetch_resource("labour_transfers.xlsx", labour_transfers_url)
    workbook: openpyxl.Workbook = openpyxl.load_workbook(path, read_only=True)
    assert set(workbook.sheetnames) == set(APP_LABOUR_SHEETS)
    for sheet in LABOUR_SHEETS:
        for row in h.parse_xlsx_sheet(
            context,
            sheet=workbook[sheet],
            skiprows=1,
        ):
            name_en = row.pop("supplier_machine_translated_english")
            if not name_en:
                context.log.info("Skipping empty row", row=row)
                continue
            addr_xuar = row.pop("address_in_xuar", None)
            addr_xuar_en = row.pop("address_in_xuar_machine_translated_english", None)
            addr = row.pop("address", None)
            addr_en = row.pop("address_machine_translated_english", None)
            sector = row.pop("industry")

            entity = context.make("Organization")
            # Setting the ID based on the translated name to avoid blanks
            entity.id = context.make_id(name_en, addr, sector)
            entity.add("name", row.pop("supplier"), lang="zhu")
            entity.add("name", name_en, lang="eng")
            entity.add("sector", sector, lang="eng")
            entity.add("keywords", row.pop("allegation", None))
            entity.add("notes", row.pop("notes"))
            entity.add("classification", sheet)
            entity.add("country", "cn")
            entity.add("topics", "export.risk")
            entity.add(
                "program",
                "Companies Named in Media and Academic Reports as engaging in Labour Transfers or other XUAR Government Programs",
            )

            row.pop("parent_company")
            parent_en = row.pop("parent_company_english")
            if parent_en is not None:
                parent_res = context.lookup("parent_company", parent_en)
                if parent_res is not None:
                    for parent_name in parent_res.values:
                        parent = context.make("Organization")
                        parent.id = context.make_id(parent_name, parent_en)
                        parent.add(
                            "name", parent_name, lang="eng", original_value=parent_en
                        )
                        parent.add("topics", "export.risk")
                        parent.add("country", parent_res.country or "cn")
                        context.emit(parent)

                        own = context.make("Ownership")
                        own.id = context.make_id("ownership", entity.id, parent.id)
                        own.add("owner", parent)
                        own.add("asset", entity)
                        context.emit(own)

                        if parent_res.ultimate:
                            ultimate = context.make("Organization")
                            ultimate.id = context.make_id(parent_res.ultimate)
                            ultimate.add("name", parent_res.ultimate, lang="eng")
                            ultimate.add("topics", "export.risk")
                            # ultimate.add("country", "cn")
                            context.emit(ultimate)

                            own = context.make("Ownership")
                            own.id = context.make_id(
                                "ownership", parent.id, ultimate.id
                            )
                            own.add("owner", ultimate)
                            own.add("asset", parent)
                            context.emit(own)
                else:
                    context.log.warning(
                        "No parent company lookup",
                        name_en=name_en,
                        parent_en=parent_en,
                    )

            apply_addresses(
                context,
                entity,
                addr=addr_xuar or addr,
                addr_en=addr_xuar_en or addr_en,
                city=None,
                city_en=None,
            )

            context.emit(entity)

        context.audit_data(
            row,
            ignore=[
                "add_date",
                "source_1",
                "updated",
                "source",
            ],
        )


def crawl_operating(context: Context, companies_url):
    path = context.fetch_resource("companies_registry.xlsx", companies_url)
    workbook: openpyxl.Workbook = openpyxl.load_workbook(path, read_only=True)
    assert set(workbook.sheetnames) == OPERATING_SHEETS
    for row in h.parse_xlsx_sheet(
        context, sheet=workbook["1. Companies Operating in XUAR"], skiprows=1
    ):
        name_en = row.pop("company_machine_translated_english")
        addr = row.pop("address")
        sector = row.pop("sector")

        entity = context.make("Organization")
        entity.id = context.make_id(name_en, addr, sector)
        entity.add("name", row.pop("company"), lang="zhu")
        entity.add("name", name_en, lang="eng")
        entity.add("sector", sector, lang="zhu")
        entity.add("sector", row.pop("sector_english"), lang="eng")
        entity.add("program", "Companies operating in the Uyghur region")
        entity.add("country", "cn")

        apply_addresses(
            context,
            entity,
            addr=addr,
            addr_en=row.pop("address_machine_translated_english"),
            city=row.pop("city"),
            city_en=row.pop("city_english"),
        )
        context.emit(entity)

    context.audit_data(row)


def crawl(context: Context):
    crawl_operating(context, OPERATING_URL)
    crawl_labour_transfers(context, LABOUR_TRANSFERS_URL)
