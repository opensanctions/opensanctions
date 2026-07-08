from openpyxl import load_workbook
from followthemoney.util import sanitize_text
from rigour.mime.types import XLSX

from zavod import Context
from zavod import helpers as h
from zavod.entity import Entity
from zavod.stateful.positions import PositionCategorisation, categorise

MISSING_ARABIC_NAME = "Said Outghilast"


def crawl_member(
    context: Context,
    row: dict[str, str | None],
    name_ara: str | None,
    position: Entity,
    categorisation: PositionCategorisation,
) -> None:
    name = row.get("prenom_et_nom")
    if name is None:
        return

    person = context.make("Person")
    # The source has no stable per-deputy id; key on name and electoral district.
    constituency = row.get("nom_de_la_circonscription")
    person.id = context.make_id(name, constituency)
    h.apply_name(person, full=name, lang="fra")
    h.apply_name(person, full=name_ara, lang="ara")
    person.add("gender", row.get("genre"))
    person.add("political", row.get("appartenance_politique"), lang="fra")
    # The right to be elected to the House of Representatives is reserved to Moroccan
    # citizens (Constitution art. 30).
    # https://www.constituteproject.org/constitution/Morocco_2011
    person.add("citizenship", "ma")

    occupancy = h.make_occupancy(
        context, person, position, categorisation=categorisation
    )
    if occupancy is None:
        return
    occupancy.add("constituency", constituency, lang="fra")
    occupancy.add(
        "politicalGroup", row.get("groupe_ou_groupement_parlementaire"), lang="fra"
    )

    context.emit(occupancy)
    context.emit(person)


def crawl(context: Context) -> None:
    path = context.fetch_resource("deputies.xlsx", context.data_url)
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    position = h.make_position(
        context,
        name="Member of the House of Representatives of Morocco",
        country="ma",
        topics=["gov.national", "gov.legislative"],
        wikidata_id="Q21328583",
        lang="eng",
    )
    categorisation = categorise(context, position)
    context.emit(position)

    workbook = load_workbook(path, read_only=True)
    arabic_names = iter(
        name
        for row in workbook["ar"].iter_rows(min_row=2, max_col=5)
        if (name := sanitize_text(row[4].value)) is not None
    )
    for row in h.parse_xlsx_sheet(context, workbook["fr"]):
        # The sheets have the same ordering, but Said Outghilast is absent from the
        # Arabic sheet. Do not advance the Arabic-name iterator for that member.
        name_ara = None
        if row.get("prenom_et_nom") != MISSING_ARABIC_NAME:
            name_ara = next(arabic_names)
        crawl_member(context, row, name_ara, position, categorisation)
