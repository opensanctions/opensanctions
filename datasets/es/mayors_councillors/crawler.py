from rigour.mime.types import XLSX
from openpyxl import load_workbook
from typing import Dict, Optional

from zavod import Context, helpers as h
from zavod.stateful.positions import categorise

IGNORE = [
    "autonomous_community",
    "column_0",
    "column_1",
    "column_2",
    "column_11",
    "column_12",
    "column_13",
    "column_14",
    "column_15",
    "column_16",
    "column_17",
]
DEFAULT_TOPICS = ["gov.muni"]
MAYOR_TOPICS = ["gov.muni", "gov.head"]
COUNCILLOR_TOPICS = ["gov.muni", "gov.legislative"]


def crawl_item(
    context: Context,
    row: Dict[str, str | None],
    period_start: str,
    period_end: Optional[str],
) -> None:
    start_date = row.pop("start_date")
    end_date = row.pop("end_date")
    assert start_date is not None

    name = row.pop("name")
    province = row.pop("province")
    municipality = row.pop("municipality")
    first_last_name = row.pop("last_name", None)
    second_last_name = row.pop("second_last_name", None)
    position_name = row.pop("position", None)

    if not name or not province or not municipality:
        context.log.warning("Missing required fields", row=row)
        return

    pep = context.make("Person")
    pep.id = context.make_id(
        name, province, municipality, first_last_name, second_last_name
    )
    # municipal elections are open to other nationals
    # https://www.senado.es/web/conocersenado/normas/constitucion/detalleconstitucioncompleta/index.html?lang=en
    pep.add("country", "es")
    if first_last_name or second_last_name:
        last_name = f"{first_last_name} {second_last_name}"
        h.apply_name(pep, first_name=name, last_name=last_name)
    else:
        pep.add("name", name)
    pep.add("political", row.pop("party"))
    pep.add("topics", "role.pep")
    # Positions are available for the current officials; historical data lists only mayors
    if not position_name:
        position_name, topics = f"Mayor of {municipality}, {province}", MAYOR_TOPICS
    else:
        translated = context.lookup_value("positions", position_name.strip())
        if translated:
            position_name = f"{translated} of {municipality}, {province}"
            topics = (
                MAYOR_TOPICS
                if "Mayor" in translated
                else COUNCILLOR_TOPICS if "Councillor" in translated else DEFAULT_TOPICS
            )
        else:
            context.log.warning("Unknown position", position=position_name.strip())
            position_name, topics = (
                f"{position_name.strip()} of {municipality}, {province}",
                DEFAULT_TOPICS,
            )
    pep.add("position", position_name)
    position = h.make_position(
        context,
        position_name,
        country="es",
        subnational_area=row.pop("ine_code"),
        lang="spa",
        topics=topics,
    )
    categorisation = categorise(context, position, is_pep=True)

    occupancy = h.make_occupancy(
        context,
        pep,
        position,
        start_date=start_date,
        end_date=end_date,
        categorisation=categorisation,
    )

    if occupancy:
        occupancy.add("constituency", province)
        occupancy.add("periodStart", period_start)
        occupancy.add("periodEnd", period_end)
        context.emit(occupancy)
        context.emit(position)
        context.emit(pep)

    context.audit_data(row, IGNORE)


def process_excel(
    context: Context,
    filename: str,
    url: str,
    title: str,
    skiprows: int,
    period_start: str,
    period_end: Optional[str],
) -> None:
    path = context.fetch_resource(filename, url)
    context.export_resource(path, XLSX, title=title)
    workbook = load_workbook(path, read_only=True)
    if len(workbook.sheetnames) != 1:
        context.log.warn(
            "Expected only one sheet in the workbook, found multiple",
            sheetnames=workbook.sheetnames,
        )
    sheet = workbook[workbook.sheetnames[0]]
    for row in h.parse_xlsx_sheet(
        context,
        sheet,
        skiprows=skiprows,
        header_lookup=context.get_lookup("columns"),
    ):
        crawl_item(context, row, period_start=period_start, period_end=period_end)


def crawl(context: Context) -> None:
    """Crawl Spanish mayors and councillors data from official sources."""
    assert context.dataset.model.url is not None
    # Process historical mayors data (2019-2023)
    hist_doc = context.fetch_html(
        context.dataset.model.url, cache_days=1, absolute_links=True
    )
    # Note: Once hist_url xpath fails on 2019_2023, check the dates in the process_excel calls below and update if needed
    hist_url = h.xpath_string(
        hist_doc,
        ".//div[@class='dnt-link-default']/a[contains(@href, 'Alcaldes_Mandato_2019_2023')]/@href",
    )
    process_excel(
        context=context,
        filename="historical.xlsx",
        url=hist_url,
        title="Mayors 2019-2023",
        skiprows=7,
        period_start="2019",
        period_end="2023",
    )
    # Process current mayors and councillors data
    current_doc = context.fetch_html(
        context.data_url, cache_days=1, absolute_links=True
    )
    current_url = h.xpath_string(
        current_doc, ".//a[@id='concejales_legislatura']/@href"
    )
    process_excel(
        context=context,
        filename="current.xlsx",
        url=current_url,
        title=context.SOURCE_TITLE,
        skiprows=5,
        period_start="2023",
        period_end="2027",
    )
