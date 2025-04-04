import re
from typing import Dict
from openpyxl import load_workbook
from rigour.mime.types import XLSX
from rigour.names import pick_name

from zavod import Context, helpers as h


# detect anything more complex than word/word/word (and handle question mark woopsie)
REGEX_LAST_NAME = re.compile(r"^[\w\?]+( ?/\s*[\w\?]+)*$")


def crawl_item(context: Context, row: Dict[str, str]):
    # Jméno fyzické osoby
    # -> First name of the natural person

    first_name_field = (row.pop("jmeno_fyzicke_osoby") or "").strip('"').strip()
    # Cleaning here rather than via type.string lookup because we assemble full
    # names from these.
    first_names = h.multi_split(first_name_field, ["/"])

    # Příjmení fyzické osoby / Název právnické osoby / Označení nebo název entity
    # -> Last name of the natural person / Name of the legal entity / Designation or name of the entity
    name_field = row.pop(
        "prijmeni_fyzicke_osoby_nazev_pravnicke_osoby_oznaceni_nebo_nazev_entity"
    )
    name_field = str(name_field)

    # Datum narození fyzické osoby
    # -> Date of birth of the natural person
    birth_date = row.pop("datum_narozeni_fyzicke_osoby")

    # Státní příslušnost fyzické osoby / sídlo právnické osoby
    # -> Nationality of the natural person / registered office of the legal entity
    countries = row.pop("statni_prislusnost_fyzicke_osoby_sidlo_pravnicke_osoby")

    sanction_props = dict()
    if REGEX_LAST_NAME.match(name_field):
        names = h.multi_split(name_field, ["/"])
    else:
        res = context.lookup("name_notes", name_field)
        if res:
            names = res.names
            sanction_props = res.sanction_props
        else:
            names = h.multi_split(name_field, ["/"])
            context.log.warning("Name field needs manual cleaning", name=name_field)

    if len(first_name_field) == 0:
        entity = context.make("LegalEntity")
        entity.id = context.make_id(name_field, first_name_field, countries)
        # There can be multiple names which are separated by /
        entity.add("name", names)
        entity.add("country", countries.split(", "), lang="ces")
    else:
        entity = context.make("Person")
        entity.id = context.make_id(name_field, first_name_field, countries)
        # There can be multiple names which are separated by /
        entity.add("lastName", names)
        entity.add("firstName", first_names)
        h.apply_name(
            entity,
            first_name=pick_name(first_names),
            last_name=pick_name(names),
        )
        h.apply_date(entity, "birthDate", birth_date.strip())
        entity.add("nationality", countries.split(", "), lang="ces")

    # Další identifikační údaje
    # -> Other identification data
    entity.add("notes", row.pop("dalsi_identifikacni_udaje"), lang="ces")

    sanction = h.make_sanction(context, entity)
    h.apply_date(sanction, "startDate", row.pop("datum_zapisu"))
    for prop, value in sanction_props.items():
        if "date" in prop.lower():
            if isinstance(value, str):
                h.apply_date(sanction, prop, value)
            elif isinstance(value, list):
                h.apply_dates(sanction, prop, value)
            else:
                raise Exception(
                    f"Unexpected value type {type(value)} for {prop}: {value}"
                )
        else:
            sanction.add(prop, value, lang="ces")

    if h.is_active(sanction):
        entity.add("topics", "sanction")

    # Popis postižitelného jednání -> Description of punishable conduct
    sanction.add("reason", row.pop("popis_postizitelneho_jednani"), lang="ces")

    # Omezující opatření -> Restrictive measures
    sanction.add("provisions", row.pop("omezujici_opatreni"), lang="ces")

    # Číslo usnesení vlády
    # -> Government resolution number
    sanction.add("recordId", row.pop("cislo_usneseni_vlady"), lang="ces")

    # Ustanovení předpisu Evropské unie, jehož skutkovou podstatu subjekt jednáním naplnil
    # -> Provision of the European Union regulation, the factual basis of which the subject fulfilled by action
    provision = row.pop(
        "ustanoveni_predpisu_evropske_unie_jehoz_skutkovou_podstatu_subjekt_jednanim_naplnil"
    )
    sanction.add("program", provision, lang="ces")

    context.emit(entity)
    context.emit(sanction)

    context.audit_data(row)


def crawl_data_url(context: Context):
    doc = context.fetch_html(context.data_url)
    # (etree.tostring(doc))
    doc.make_links_absolute(context.data_url)
    anchor = doc.xpath('//a/span[text()="Vnitrostátní sankční seznam"]/..')
    assert len(anchor) != 1, len(anchor)
    return anchor[0].get("href")


def crawl(context: Context) -> None:
    # First we find the link to the excel file
    data_url = crawl_data_url(context)
    path = context.fetch_resource("source.xlsx", data_url)
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    wb = load_workbook(path, read_only=True)
    if len(wb.sheetnames) != 1:
        raise Exception("Expected only one sheet in the workbook")

    for row in h.parse_xlsx_sheet(context, wb[wb.sheetnames[0]]):
        crawl_item(context, row)
