import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from rigour.mime.types import XLSX

from zavod import Context, helpers as h


def get_most_recent_link(context, doc):
    # Select all .xlsx resource links within resource-item elements
    links = doc.xpath(
        "//li[contains(@class, 'resource-item')]//a[contains(@class, 'resource-url-analytics') and contains(@href, '.xlsx')]"
    )
    dated_links = []

    for link in links:
        href = link.get("href")
        # The XPath selects all downloadable XLSX links on the page, including outdated ones.
        # However, we're only interested in the most recent file, so we filter out known outdated links,
        # such as one with a hardcoded 2024 date and any from 2019.
        # These older links follow a different URL format and lack consistent date patterns.
        # Since we assert that the latest link must be within the last 30 days,
        # it's safe and more efficient to exclude them upfront to simplify parsing.
        if "17.06.2024" in href or re.search(r"resources/2019-\d{2}", href):
            continue
        # Full date: YYYY.MM.DD
        match = re.search(r"(\d{4})[.](\d{2})[.](\d{2})", href)
        if match:
            date_obj = datetime.strptime(match.group(0), "%Y.%m.%d")
            dated_links.append((date_obj, href))
        else:
            context.log.warning(f"Link {href} does not contain a date")
    # Select the most recent dated link
    latest_link = max(dated_links, key=lambda x: x[0])
    # Ensure the latest file is from the last 30 days
    assert datetime.now() - timedelta(days=30) < latest_link[0]

    return latest_link[1]


def crawl_row(context, row):
    tax_number = row.pop("tax_number")
    name = row.pop("name")
    director = row.pop("director")
    entity = context.make("Organization")
    entity.id = context.make_id(tax_number, name)
    entity.add("name", name)
    entity.add("legalForm", row.pop("legal_form"))
    entity.add("country", "md")
    entity.add("taxNumber", tax_number)
    address = h.make_address(
        context,
        full=row.pop("address"),
        place=row.pop("admin_unit_code"),
    )
    h.copy_address(entity, address)
    h.apply_date(entity, "incorporationDate", row.pop("incorporation_date"))
    h.apply_date(entity, "dissolutionDate", row.pop("dissolution_date"))

    if director:
        dir = context.make("Person")
        dir.id = context.make_id(director)
        dir.add("name", director)

        directorship = context.make("Directorship")
        directorship.id = context.make_id(entity.id, dir.id)
        directorship.add("organization", entity.id)
        directorship.add("director", dir.id)

        context.emit(dir)
        context.emit(directorship)

    context.emit(entity)
    context.audit_data(row)


def crawl(context: Context):
    doc = context.fetch_html(context.data_url, cache_days=1)
    data_url = get_most_recent_link(context, doc)
    path = context.fetch_resource("list.xlsx", data_url)
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    wb = load_workbook(path, read_only=True)
    assert set(wb.sheetnames) == {wb.active.title}

    for row in h.parse_xlsx_sheet(
        context, wb["organizations"], skiprows=4, header_lookup="columns"
    ):
        crawl_row(context, row)
