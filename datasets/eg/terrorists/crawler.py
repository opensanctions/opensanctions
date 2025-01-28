from typing import List

from normality import collapse_spaces
from openpyxl import load_workbook
from pantomime.types import XLSX

from zavod import Context, helpers as h
from zavod.shed.zyte_api import fetch_html, fetch_resource


DATE_SPLITS = [
    " arab ",
    " مد القرار",  # Decision extended
    " إعادة النشر في",  # Republished in
    "المد في ",  # extended in
    "إعادة المد في",  # re-extend in
    "إعادة",  # re-
    "إعادة إدراج في",  # re-insert in
    "نشر",  # publish
    "إدراج في",  # insert
]


def arabic_to_latin(arabic_date):
    arabic_numerals = "٠١٢٣٤٥٦٧٨٩"
    latin_numerals = "0123456789"
    transtable = str.maketrans(arabic_numerals, latin_numerals)
    return arabic_date.translate(transtable)


def clean_date(date_str: str) -> List[str]:
    if not date_str:
        return []
    latin_date = arabic_to_latin(date_str)
    return h.multi_split(collapse_spaces(latin_date), DATE_SPLITS)


def crawl_terrorist(input_dict: dict, context: Context):
    name = input_dict.pop("name")
    case_number = input_dict.pop("case_number")

    person = context.make("Person")
    person.id = context.make_id(name, case_number)
    person.add("name", name)
    person.add("alias", input_dict.pop("alias"))
    person.add("nationality", input_dict.pop("nationality"), lang="ara")
    person.add("country", "eg")
    person.add("passportNumber", input_dict.pop("passport"))
    person.add("idNumber", input_dict.pop("national_id"))
    person.add("topics", "sanction.counter")

    sanction = h.make_sanction(context, person, case_number)
    h.apply_dates(
        sanction, "listingDate", clean_date(input_dict.pop("date_of_publication"))
    )
    sanction.add("description", f"Case number: {case_number}")
    sanction.add("authorityId", input_dict.pop("terrorist_designation_decision_number"))
    sanction.add(
        "description",
        f"Publication Page: {input_dict.pop('number_of_publication')}",
    )

    context.emit(person, target=True)
    context.emit(sanction)
    context.audit_data(input_dict)


def crawl_terrorist_entities(input_dict: dict, context: Context):
    name = input_dict.pop("name")
    case_number = input_dict.pop("case_number")

    if not name:
        return

    entity = context.make("LegalEntity")
    entity.id = context.make_id(name, case_number)
    entity.add("name", name)
    entity.add("country", "eg")
    entity.add("topics", "sanction.counter")

    gazette_issue = input_dict.pop("issue_in_official_gazette")
    sanction = h.make_sanction(context, entity, case_number + gazette_issue)
    sanction.add("description", f"Case number: {case_number}")
    sanction.add("description", f"Issue in official gazette: {gazette_issue}")
    sanction.add("summary", input_dict.pop("updates"), lang="ara")

    context.emit(entity, target=True)
    context.emit(sanction)
    context.audit_data(input_dict, ignore=["series"])


def crawl_legal_persons(input_dict: dict, context: Context):
    name = input_dict.pop("name")
    case_number = input_dict.pop("case_number")

    if not name:
        return

    entity = context.make("LegalEntity")
    entity.id = context.make_id(name, case_number)
    entity.add("name", name)
    entity.add("country", "eg")

    entity.add("address", input_dict.pop("headquarters"))
    entity.add("registrationNumber", input_dict.pop("commercial_registration_number"))
    entity.add("topics", "sanction.counter")

    gazette_issue = input_dict.pop("issue_in_official_gazette")
    sanction = h.make_sanction(context, entity, case_number + gazette_issue)
    sanction.add("description", f"Case number: {case_number}")
    sanction.add("description", f"Issue in official gazette: {gazette_issue}")
    sanction.add("summary", input_dict.pop("updates"), lang="ara")

    context.emit(entity, target=True)
    context.emit(sanction)
    context.audit_data(input_dict, ignore=["series"])


def crawl(context: Context):
    response = fetch_html(
        context,
        context.data_url,
        ".//*[@class='LinkStyle AutoDownload']",
        geolocation="eg",
    )
    response.make_links_absolute(context.data_url)

    excel_link = response.find(".//*[@class='LinkStyle AutoDownload']").get("href")

    _, _, _, path = fetch_resource(
        context, "list.xlsx", excel_link, XLSX, geolocation="eg"
    )
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    wb = load_workbook(path, read_only=True)

    for item in h.parse_xlsx_sheet(context, wb["الإرهابيين"], header_lookup="columns"):
        crawl_terrorist(item, context)

    for item in h.parse_xlsx_sheet(
        context, wb["الكيانات الإرهابية"], skiprows=1, header_lookup="columns"
    ):
        crawl_terrorist_entities(item, context)

    for item in h.parse_xlsx_sheet(
        context, wb["الشخصيات الاعتبارية"], skiprows=1, header_lookup="columns"
    ):
        crawl_legal_persons(item, context)

    assert len(wb.sheetnames) == 3, wb.sheetnames
