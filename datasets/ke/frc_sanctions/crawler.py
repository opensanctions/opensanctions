from typing import Dict

from normality import stringify
from openpyxl import load_workbook
from rigour.mime.types import XLSX

from zavod import Context
from zavod import helpers as h


def parse_aliases(raw: str | None) -> list[str]:
    """Split newline-separated numbered aliases: "1. FOO\n2. BAR" → ["FOO", "BAR"]."""
    if not raw:
        return []
    results = []
    for line in raw.splitlines():
        line = line.strip()
        if not line or line.lower() == "n/a":
            continue
        # Strip leading numbering "1. ", "2. " etc.
        if line[0].isdigit() and ". " in line:
            line = line.split(". ", 1)[1]
        results.append(line)
    return results


def crawl_row(context: Context, row: Dict[str, str | None]) -> None:
    # parse_xlsx_sheet slugifies headers: "Full Name" -> "full_name"
    # Pre-clean n/a placeholders so entity.add() receives None rather than the string
    row = {
        k: (None if isinstance(v, str) and v.strip().lower() == "n/a" else v)
        for k, v in row.items()
    }
    reference = row.pop("reference")
    category = row.pop("category")
    assert category == "Individual", f"Unexpected category: {category!r}"

    person = context.make("Person")
    person.id = context.make_slug(reference)

    name = row.pop("full_name")
    assert name, f"Missing name for {reference}"
    person.add("name", name)
    person.add("title", stringify(row.pop("title")))
    for alias in parse_aliases(row.pop("aliases", None)):
        person.add("alias", alias)

    person.add("idNumber", stringify(row.pop("id_number")))
    person.add("passportNumber", stringify(row.pop("passport_number")))
    person.add("gender", row.pop("gender"))
    h.apply_date(person, "birthDate", row.pop("date_of_birth"))
    h.apply_date(person, "birthDate", row.pop("alternative_date_of_birth"))
    person.add("birthPlace", stringify(row.pop("place_of_birth")))
    person.add("nationality", stringify(row.pop("nationality_1")))
    person.add("nationality", stringify(row.pop("nationality_2")))
    person.add("address", stringify(row.pop("physical_address")))
    person.add("position", stringify(row.pop("occupation")))
    person.add("phone", stringify(row.pop("telephone_number")))
    person.add("topics", "sanction")

    sanction = h.make_sanction(context, person)
    h.apply_date(sanction, "listingDate", row.pop("date_of_designation"))
    h.apply_date(sanction, "modifiedAt", row.pop("last_update_on"))
    sanction.add("reason", stringify(row.pop("narrative_summary")))

    context.emit(person)
    context.emit(sanction)
    context.audit_data(row, ignore=["postal_address", "column_0"])


def crawl(context: Context) -> None:
    # Step 1: find the current XLSX URL from the listing page
    doc = context.fetch_html(context.data_url, absolute_links=True)
    xlsx_links = h.xpath_strings(doc, ".//a[contains(@href, 'Domestic-List')]/@href")
    if not xlsx_links:
        xlsx_links = h.xpath_strings(doc, ".//a[contains(@href, '.xlsx')]/@href")
    assert len(xlsx_links) > 0, "Could not find XLSX link on FRC domestic list page"
    xlsx_url = xlsx_links[0]

    # Step 2: download and parse
    path = context.fetch_resource("source.xlsx", xlsx_url)
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    wb = load_workbook(path, read_only=True)
    assert "Sheet1" in wb.sheetnames, wb.sheetnames
    for row in h.parse_xlsx_sheet(context, wb["Sheet1"]):
        crawl_row(context, row)
