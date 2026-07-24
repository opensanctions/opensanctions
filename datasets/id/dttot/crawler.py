import re
import openpyxl
from normality import squash_spaces
from rigour.mime.types import XLSX

from zavod import Context, helpers as h
from zavod.extract import zyte_api

# Date range like "00/00/1970-1973"
DATE_RANGE_RE = re.compile(r"00/00/(\d{4})-(\d{4})")
ADDR_DELIM = re.compile(r"[\W][a-zA-Z]\)|;")
# eng: 'Branch office 4:'
BRANCH_PATTERN = re.compile(r"^kantor cabang \d+[,:]\s*", re.IGNORECASE)


def clean_addresses(raw_addresses: str) -> list[str]:
    """Split and clean a multi-address string into individual addresses."""
    cleaned = []
    for address_block in h.multi_split(
        raw_addresses, ["\n-\t", "\n- ", ", - ", ", -,"]
    ):
        for address in ADDR_DELIM.split(address_block):
            address = squash_spaces(BRANCH_PATTERN.sub("", address).strip("-"))
            if address and not BRANCH_PATTERN.match(address):
                cleaned.append(address)
    return cleaned


def crawl_row(context: Context, row: dict[str, str | None]) -> None:
    item_id = row.pop("id")
    res = context.lookup("type", row.pop("type"), warn_unmatched=True)
    assert res and res.value
    entity = context.make(res.value)
    names = h.multi_split(row.pop("name"), ["alias", "ALIAS", "Alias"])
    entity.id = context.make_id(item_id, *names)
    entity.add("name", names[0])
    entity.add("alias", names[1:])
    entity.add("topics", "sanction")
    for c in h.multi_split(row.pop("country"), ["\n"]):
        entity.add("country", c.strip("-"))
    if notes := row.pop("description"):
        entity.add("notes", squash_spaces(notes))
    if raw_addresses := row.pop("address"):
        for address in clean_addresses(raw_addresses):
            entity.add("address", address)

    dob_raw = row.pop("birth_date")
    birth_place = row.pop("birth_place")
    if entity.schema.is_a("Person"):
        entity.add("birthPlace", birth_place)
        for dob in h.multi_split(dob_raw, [" atau ", "\n", "- ", " / "]):
            if match := DATE_RANGE_RE.match(dob):
                # Date range like "00/00/1970-1973" - add both years
                h.apply_dates(entity, "birthDate", list(match.groups()))
            else:
                h.apply_date(entity, "birthDate", dob)

    sanction = h.make_sanction(context, entity)
    sanction.add("authorityId", item_id)

    context.emit(entity)
    context.emit(sanction)

    context.audit_data(row)


def crawl(context: Context) -> None:
    # The source page wraps this single download link in a markup structure that
    # changes from time to time, so select the unique .xlsx anchor directly rather
    # than depending on a specific ancestor element. xpath_string still asserts a
    # single match, so any future duplication crashes loudly.
    xlsx_link_xpath = ".//a[contains(@href, '.xlsx')]/@href"
    doc = zyte_api.fetch_html(
        context,
        context.data_url,
        unblock_validator=xlsx_link_xpath,
        geolocation="ID",
    )
    xlsx_link = h.xpath_string(doc, xlsx_link_xpath)
    _, _, _, path = zyte_api.fetch_resource(
        context,
        "source.xlsx",
        xlsx_link,
        expected_media_type=XLSX,
        geolocation="ID",
    )
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)
    wb: openpyxl.Workbook = openpyxl.load_workbook(path, read_only=True)
    assert len(wb.sheetnames) == 1, wb.sheetnames
    for row in h.parse_xlsx_sheet(
        context, wb["Export"], header_lookup=context.get_lookup("headers")
    ):
        crawl_row(context, row)
