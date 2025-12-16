from pathlib import Path
from zavod import Context
from zavod import helpers as h
from lxml.html import HtmlElement
from openpyxl import load_workbook
from rigour.mime.types import XLSX

AUDIT_FIELDS = [
    "row_no",
    "mother_name",
    "birth_place",
    "first_and_surname",
    "country",
    "doc_type",
    "doc_no",
    "doc_issue_auth",
    "doc_issue_date",
    "description",
]


class MissingDataFile(Exception):
    pass


def xlsx_check(doc: HtmlElement) -> bool | str:
    for link in doc.iterlinks():
        ext = link[2].split(".")[-1]
        if "xlsx" in ext:
            return str(link[2])
    return False


def crawl_person(context: Context, path: Path) -> None:
    wb = load_workbook(path, read_only=True)
    for row in h.parse_xlsx_sheet(
        context,
        wb.worksheets[0],
        skiprows=2,
        header_lookup=context.get_lookup("columns"),
    ):
        national_id = row.pop("national_id")
        fullname_en = row.pop("full_name_en")
        fullname_ar = row.pop("full_name_ar")
        birth_date = row.pop("birth_date")
        included_date = row.pop("included_date")
        country = row.pop("country")

        person = context.make("Person")
        person.id = context.make_id(
            national_id,
            fullname_ar,
            birth_date,
        )

        person.add("name", fullname_ar, lang="ara")
        person.add("name", fullname_en, lang="eng")
        person.add("country", country)
        person.add("idNumber", national_id)
        person.add("nationality", row.pop("nationality"))
        h.apply_date(person, "birthDate", birth_date)
        person.add("sourceUrl", context.data.url)

        address_ent = h.make_address(
            context,
            street=row.pop("area_and_street"),
            city=row.pop("city"),
            country=country,
            lang="ara",
        )
        h.copy_address(person, address_ent)

        sanction = h.make_sanction(context, person)
        h.apply_date(sanction, "startDate", included_date)

        context.emit(person)
        context.emit(sanction)

        context.audit_data(row, AUDIT_FIELDS)


def crawl_legal_entities(context: Context, path: Path) -> None:
    wb = load_workbook(path, read_only=True)
    for row in h.parse_xlsx_sheet(
        context,
        wb.worksheets[1],
        skiprows=1,
        header_lookup=context.get_lookup("columns"),
    ):
        full_name_en = row.pop("full_name_en")
        included_date = row.pop("included_date")

        legalent = context.make("LegalEntity")
        legalent.id = context.make_id(full_name_en)

        legalent.add("name", row.pop("full_name_ar"), lang="ara")
        legalent.add("name", full_name_en, lang="eng")
        legalent.add("classification", row.pop("classification"), lang="ara")

        sanction = h.make_sanction(context, legalent)
        h.apply_date(sanction, "startDate", included_date)

        context.emit(legalent)
        context.emit(sanction)

        context.audit_data(row, AUDIT_FIELDS)


def crawl(context: Context) -> None:
    doc = context.fetch_html(context.data_url, cache_days=5)
    url = xlsx_check(doc)
    if url is False:
        msg = "XLSX file does not exist!"
        context.log.error(msg)
        raise MissingDataFile(msg)

    path = context.fetch_resource("lists.xlsx", url)
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    crawl_person(context, path)
    crawl_legal_entities(context, path)
