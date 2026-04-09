from openpyxl import load_workbook
from typing import Dict, Any
from urllib.parse import urljoin
from rigour.mime.types import XLSX

from zavod import Context, helpers as h

BASE_ACTION_URL = "https://reportdocs.static.szse.cn/UpFiles/zqjghj/"


def crawl_row(context: Context, row: Dict[str, Any]) -> None:
    id = row.pop("company_code")
    # e.g. 'Sunflower' instead of 'Sunflower Health Technology Co., Ltd.'
    short_name = row.pop("company_abbreviation")

    # skip empty company names (44 cases)
    if short_name is None:
        context.log.info("Skipping row with empty company name", row=row)
        return

    entity = context.make("LegalEntity")
    entity.id = context.make_id(id, short_name)
    entity.add("name", short_name)
    entity.add("country", "cn")
    entity.add("topics", "reg.action")

    sanction = h.make_sanction(context, entity)
    h.apply_date(sanction, "listingDate", row.pop("date_of_action"))
    sanction.add("sourceUrl", urljoin(BASE_ACTION_URL, row.pop("action_pdf")))
    sanction.add("summary", row.pop("decision_summary"))
    sanction.add("provisions", row.pop("action_type"))

    context.emit(entity)
    context.emit(sanction)

    context.audit_data(row)


def crawl(context: Context) -> None:
    path = context.fetch_resource("source.xlsx", context.data_url)
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    wb = load_workbook(path, read_only=False)
    # disciplinary actions
    assert set(wb.sheetnames) == {"纪律处分"}, wb.sheetnames

    for row in h.parse_xlsx_sheet(
        context,
        wb[wb.sheetnames[0]],
        header_lookup=context.get_lookup("columns"),
    ):
        crawl_row(context, row)
