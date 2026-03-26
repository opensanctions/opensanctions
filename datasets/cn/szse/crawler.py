from openpyxl import load_workbook
from typing import Dict, Any

from zavod import Context, helpers as h
from rigour.mime.types import XLSX


MEASURES_API_URL = "https://www.szse.cn/api/report/ShowReport?SHOWTYPE=xlsx&CATALOGID=1800_jgxxgk&TABKEY=tab1&selectBkmc=0"
ACTIONS_API_URL = "https://www.szse.cn/api/report/ShowReport?SHOWTYPE=xlsx&CATALOGID=1800_jgxxgk_cf&TABKEY=tab2&selectGsbk=0"


def fetch_xlsx_rows(context: Context, api_url: str, filename: str, referer_url: str):
    path = context.fetch_resource(filename, api_url, headers={"Referer": referer_url})
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)
    wb = load_workbook(path, read_only=False)
    if len(wb.sheetnames) != 1:
        raise ValueError(f"Expected 1 sheet, got {len(wb.sheetnames)} in {filename}")
    return h.parse_xlsx_sheet(
        context, wb[wb.sheetnames[0]], header_lookup=context.get_lookup("columns")
    )


def crawl_measure_item(context: Context, row: Dict[str, Any]) -> None:
    id = row.pop("company_code")
    name = row.pop("company_name")

    entity = context.make("LegalEntity")
    entity.id = context.make_id(id, name)
    entity.add("name", name)
    entity.add("country", "cn")

    sanction = h.make_sanction(context, entity)
    h.apply_date(sanction, "startDate", row.pop("date_of_measure"))
    # source + reason:
    sanction.add(
        "sourceUrl",
        "https://reportdocs.static.szse.cn/UpFiles/zqjghj/" + row.pop("letter_pdf"),
    )

    sanction.add("description", row.pop("measure_type"))
    sanction.add("scope", row.pop("subject"))

    breakpoint()


# def crawl_actions_item(context: Context, row: Dict[str, str]) -> None:
#     print("bla bla")


def crawl(context: Context) -> None:
    doc = context.fetch_html(context.data_url, cache_days=1, absolute_links=True)

    # get urls to data
    measures_url = h.xpath_string(
        doc,
        ".//ul[@class='side-menu-con']//li[@chnlcode='superviseMeasure_report']//a/@href",
    )
    # actions_url = h.xpath_string(
    #     doc,
    #     ".//ul[@class='side-menu-con']//li[@chnlcode='disciplinaryAction_report']//a/@href",
    # )
    # restrict_trading_url = h.xpath_string(
    #     doc, ".//ul[@class='side-menu-con']//li[@chnlcode='restrict_trading']//a/@href"
    # )

    # collect regulatory measures
    for row in fetch_xlsx_rows(
        context, MEASURES_API_URL, "measures_source.xlsx", measures_url
    ):
        print(measures_url)
        crawl_measure_item(context, row)

    # collect disciplinary actions
    # for row in fetch_xlsx_rows(context, ACTIONS_API_URL, "actions_source.xlsx", actions_url):
    #     crawl_actions_item(context, row)

    # collect restricted trading
    # ... restrict_trading_url code ...
