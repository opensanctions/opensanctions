from typing import Dict, Generator
from rigour.mime.types import XLSX
from openpyxl import load_workbook
from normality import slugify, stringify
from datetime import datetime
import openpyxl

from zavod import Context, helpers as h


def parse_sheet(
    context: Context,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    skiprows: int = 0,
) -> Generator[dict, None, None]:
    headers = None

    row_counter = 0

    for row in sheet.iter_rows():
        # Increment row counter
        row_counter += 1

        # Skip the desired number of rows
        if row_counter <= skiprows:
            continue
        cells = [c.value for c in row]
        if headers is None:
            headers = []
            for idx, cell in enumerate(cells):
                if cell is None:
                    cell = f"column_{idx}"
                headers.append(slugify(cell))
            continue

        record = {}
        for header, value in zip(headers, cells):
            if isinstance(value, datetime):
                value = value.date()
            record[header] = stringify(value)
        if len(record) == 0:
            continue
        yield record


def crawl_item(row: Dict[str, str], context: Context):

    name = row.pop("provider-name")

    if not name:
        return

    entity = context.make("LegalEntity")
    entity.id = context.make_id(name)
    entity.add("name", name)
    if row.get("npi-num"):
        entity.add("notes", "NPI: " + row.pop("npi-num"))
    else:
        row.pop("npi-num")

    entity.add("topics", "sanction")

    sanction = h.make_sanction(context, entity)
    sanction.add(
        "startDate",
        h.parse_date(row.pop("exclusion-effective-date"), formats=["%m/%d/%Y"]),
    )

    context.emit(entity, target=True)
    context.emit(sanction)

    context.audit_data(row, ignore=["license-num", "provider-type"])


def crawl(context: Context) -> None:
    path = context.fetch_resource("list.xlsx", context.data_url)
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    wb = load_workbook(path, read_only=True)

    for item in parse_sheet(context, wb.active):
        crawl_item(item, context)
        