from typing import Dict, Generator
from rigour.mime.types import XLSX
from openpyxl import load_workbook
from normality import slugify, stringify
from datetime import datetime
import openpyxl

from zavod import Context, helpers as h


def parse_sheet(
    context: Context,
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    skiprows: int = 0,
) -> Generator[dict, None, None]:
    headers = None

    row_counter = 0

    for row in sheet.iter_rows():
        # Increment row counter
        row_counter += 1

        # Skip the desired number of rows
        if row_counter <= skiprows:
            continue
        cells = [c.value for c in row]
        if headers is None:
            headers = []
            for idx, cell in enumerate(cells):
                if cell is None:
                    cell = f"column_{idx}"
                headers.append(slugify(cell))
            continue

        record = {}
        for header, value in zip(headers, cells):
            if isinstance(value, datetime):
                value = value.date()
            record[header] = stringify(value)
        if len(record) == 0:
            continue
        yield record


def crawl_item(row: Dict[str, str], context: Context):

    first_name = row.pop("first-name")
    last_name = row.pop("last-name")
    business_name = row.pop("business-name")

    if not first_name and not last_name and not business_name:
        return

    if business_name:
        entity = context.make("Company")
        entity.id = context.make_id(business_name)
        entity.add("name", business_name)
    else:
        entity = context.make("Person")
        entity.id = context.make_id(first_name, last_name)
        entity.add("firstName", first_name)
        entity.add("lastName", last_name)

    if row.get("npi"):
        entity.add("npiCode", row.pop("npi"))

    entity.add("topics", "debarment")

    sanction = h.make_sanction(context, entity)
    sanction.add("startDate", h.parse_date(row.pop("sancdate"), formats=["%Y%m%d"]))

    context.emit(entity, target=True)
    context.emit(sanction)

    # general is the medical field of the professional
    context.audit_data(row, ignore=["general", "state"])


def crawl_excel_url(context: Context):
    doc = context.fetch_html(context.data_url)
    doc.make_links_absolute(context.data_url)
    return doc.find(".//a[@data-text='List of Excluded Individuals']").get("href")


def crawl(context: Context) -> None:
    # First we find the link to the excel file
    excel_url = crawl_excel_url(context)
    path = context.fetch_resource("list.xlsx", excel_url)
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    wb = load_workbook(path, read_only=True)

    for item in parse_sheet(context, wb["Sheet1"], skiprows=2):
        crawl_item(item, context)
