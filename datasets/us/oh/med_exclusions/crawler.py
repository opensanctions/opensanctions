from itertools import product
from rigour.mime.types import XLSX
from openpyxl import load_workbook
import re

from zavod import Context, helpers as h
from zavod.extract import zyte_api

REGEX_DBA = re.compile(r"\bdba\b", re.IGNORECASE)
REGEX_AKA = re.compile(r"\(?a\.?k\.?a\b\.?|\)", re.IGNORECASE)


def crawl_individual(row: dict[str, str | None], context: Context) -> None:
    entity = context.make("Person")
    entity.id = context.make_id(
        row.get("last_name"), row.get("first_name"), row.get("npi")
    )
    last_names = REGEX_AKA.split(row.pop("last_name") or "")
    middle_names = REGEX_AKA.split(row.pop("middle_name") or "")
    first_names = REGEX_AKA.split(row.pop("first_name") or "")
    for first_name, middle_name, last_name in product(
        first_names, middle_names, last_names
    ):
        h.apply_name(
            entity, first_name=first_name, middle_name=middle_name, last_name=last_name
        )

    entity.add("country", "us")
    entity.add("sector", row.pop("provider_type"))
    entity.add("topics", "debarment")
    if (provider_id := row.pop("provider_id")) is not None:
        entity.add("description", "Provider ID: " + provider_id)
    h.apply_date(entity, "birthDate", row.pop("dob"))

    if row.get("npi"):
        npis = h.multi_split(row.pop("npi"), ["N/A", "n/a", "/", "\n", " "])

        entity.add("npiCode", npis)
    else:
        row.pop("npi")

    sanction = h.make_sanction(context, entity)
    h.apply_date(sanction, "startDate", row.pop("action_date"))
    sanction.add("status", row.pop("status"))

    context.emit(entity)
    context.emit(sanction)

    context.audit_data(row, ignore=["date_added"])


def crawl_organization(row: dict[str, str | None], context: Context) -> None:
    entity = context.make("Company")
    entity.id = context.make_id(row.get("organization_name"))
    names = REGEX_DBA.split(row.pop("organization_name") or "")
    aliases = names[1:]
    names = names[0].split("Owner:")
    entity.add("name", names[0])
    entity.add("alias", aliases)
    entity.add("country", "us")
    entity.add("sector", row.pop("provider_type"))
    entity.add("topics", "debarment")
    if row.get("provider_id"):
        entity.add("description", "Provider ID: " + (row.pop("provider_id") or ""))

    for owner_name in names[1:]:
        owner = context.make("LegalEntity")
        owner.id = context.make_id(row.get("organization_name"), owner_name)
        owner.add("name", owner_name.strip())
        owner.add("country", "us")

        rel = context.make("Ownership")
        rel.id = context.make_id(entity.id, owner.id)
        rel.add("owner", owner)
        rel.add("asset", entity)

        context.emit(owner)
        context.emit(rel)

    address = h.make_address(
        context,
        street=row.pop("address_1"),
        street2=row.pop("address_2"),
        state=row.pop("state"),
        city=row.pop("city"),
        postal_code=row.pop("zip_code"),
        country_code="us",
    )
    h.apply_address(context, entity, address)

    if row.get("npi"):
        npis = h.multi_split(row.pop("npi"), ["N/A", "n/a", "/", "\n", " "])

        entity.add("npiCode", npis)
    else:
        row.pop("npi")

    sanction = h.make_sanction(context, entity)
    h.apply_date(sanction, "startDate", row.pop("action_date"))
    sanction.add("status", row.pop("status"))

    context.emit(entity)
    context.emit(sanction)
    if address is not None:
        context.emit(address)

    context.audit_data(row, ignore=["date_added"])


def crawl_excel_url(context: Context) -> str:
    file_a_xpath = ".//a[contains(text(), 'Medicaid') and contains(text(), 'Exclusion') and contains(text(), 'Suspension') and contains(@href, 'xlsx')]"
    doc = zyte_api.fetch_html(context, context.data_url, unblock_validator=file_a_xpath)
    url = h.xpath_string(doc, file_a_xpath + "/@href")
    assert url is not None, "Could not find Excel file URL"
    return url


def crawl(context: Context) -> None:
    # First we find the link to the excel file
    excel_url = crawl_excel_url(context)
    _, _, _, path = zyte_api.fetch_resource(
        context, "source.xlsx", excel_url, expected_media_type=XLSX, geolocation="US"
    )
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    wb = load_workbook(path, read_only=True)
    sheet_names = wb.sheetnames
    sheet_names.remove("Instructions")

    for item in h.parse_xlsx_sheet(context, wb["Individuals"]):
        crawl_individual(item, context)
    sheet_names.remove("Individuals")

    for item in h.parse_xlsx_sheet(context, wb["Organizations"]):
        crawl_organization(item, context)
    sheet_names.remove("Organizations")

    # Detect if new sheets show up.
    assert sheet_names == [], sheet_names
