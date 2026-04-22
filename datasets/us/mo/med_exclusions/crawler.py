from typing import Dict
from rigour.mime.types import XLSX
from openpyxl import load_workbook

from zavod import Context
from zavod.extract import zyte_api
from zavod import helpers as h


def crawl_item(row: Dict[str, str | None], context: Context) -> None:
    entity = context.make("LegalEntity")
    entity.id = context.make_id(row.get("provider_name"), row.get("npi"))
    entity.add("name", row.pop("provider_name"))
    if row.get("npi") != "N/A":
        entity.add("npiCode", row.pop("npi"))
    else:
        row.pop("npi")
    entity.add("topics", "debarment")
    entity.add("sector", row.pop("prov_type_specialty"))
    if row.get("license") and row.get("license") != "N/A":
        entity.add("description", f"License number: {row.pop('license')}")
    else:
        row.pop("license")
    entity.add("country", "us")

    sanction = h.make_sanction(context, entity)
    h.apply_date(sanction, "startDate", row.pop("term_date"))
    h.apply_date(sanction, "date", row.pop("letter_date"))
    sanction.add("reason", row.pop("termination_reason"))

    context.emit(entity)
    context.emit(sanction)

    context.audit_data(row)


def crawl_excel_url(context: Context, sanctions_list_url: str) -> str:
    excel_xpath = (
        ".//a[contains(@href, 'Sanction-List') and contains(@href, '.xlsx')]/@href"
    )
    exclusions_page = zyte_api.fetch_html(
        context,
        sanctions_list_url,
        unblock_validator=excel_xpath,
    )
    excel_url = h.xpath_string(exclusions_page, excel_xpath)
    return excel_url


def crawl(context: Context) -> None:
    # Locate the Sanctions page
    sanctions_list_url_xpath = ".//a[contains(@href, 'sanction-list')]/@href"
    landing_page = zyte_api.fetch_html(
        context,
        context.data_url,
        unblock_validator=sanctions_list_url_xpath,
    )
    sp_url = h.xpath_string(landing_page, sanctions_list_url_xpath)

    # Find the link to the excel file on the Sanctions page
    excel_url = crawl_excel_url(context, sp_url)
    _, _, _, path = zyte_api.fetch_resource(
        context, "source.xlsx", excel_url, expected_media_type=XLSX, geolocation="US"
    )
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)
    wb = load_workbook(path, read_only=True)

    if len(wb.sheetnames) > 1:
        context.log.warning("Additional sheet found")

    assert wb.active is not None, "No active sheet found in workbook"
    for item in h.parse_xlsx_sheet(context, wb.active, skiprows=1):
        crawl_item(item, context)
