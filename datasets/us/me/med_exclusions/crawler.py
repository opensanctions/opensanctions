import json
from typing import Dict

from openpyxl import load_workbook
from rigour.mime.types import XLSX
from zavod.extract import zyte_api

from zavod import Context
from zavod import helpers as h


EXCEL_FILENAME = "PI0008-PM%20Monthly%20Exclusion%20Report.xlsx"
BASE_URL = "https://mainecare.maine.gov/PrvExclRpt"


def crawl_item(row: Dict[str, str | None], context: Context) -> None:
    if first_name := row.pop("provider_first_name"):
        last_name = row.pop("provider_last_name")
        middle_initial = row.pop("provider_mi")
        position = row.pop("provider_type")

        entity = context.make("Person")
        entity.id = context.make_id(first_name, last_name, middle_initial)
        h.apply_name(
            entity,
            first_name=first_name,
            last_name=last_name,
            middle_name=middle_initial,
        )
        if position and position.lower() != "other":
            entity.add("position", position)
    else:
        last_name = row.pop("provider_last_name")
        entity = context.make("Company")
        entity.id = context.make_id(last_name)
        entity.add("name", last_name)
        entity.add("sector", row.pop("provider_type"))

    # Number of alias
    for i in [1, 2, 3, 4]:
        alias_first_name = row.pop(f"alias_first_name_{i}")
        alias_last_name = row.pop(f"alias_last_name_{i}")

        if not alias_first_name and not alias_last_name:
            continue

        # If the entity is a company and there is an alias, we consider it as an unknown link
        if entity.schema.name == "Company":
            person = context.make("Person")
            person.id = context.make_id(alias_first_name, alias_last_name)
            h.apply_name(person, first_name=alias_first_name, last_name=alias_last_name)
            person.add("country", "us")

            link = context.make("UnknownLink")
            link.id = context.make_id(person.id, entity.id)
            link.add("object", entity)
            link.add("subject", person)

            context.emit(link)
            context.emit(person)

        else:
            h.apply_name(
                entity,
                first_name=alias_first_name,
                last_name=alias_last_name,
                alias=True,
            )

    entity.add("country", "us")
    entity.add("topics", "debarment")

    sanction = h.make_sanction(context, entity)

    h.apply_date(sanction, "startDate", row.pop("state_exclusion_start_date"))

    context.emit(entity)
    context.emit(sanction)

    context.audit_data(row)


def crawl_excel_url(context: Context) -> str:
    _, _, _, txt = zyte_api.fetch_text(context, context.data_url, geolocation="us")
    # Parse out the table data JSON embedded in the HTML
    start = txt.find("WPQ2ListData") + 15
    end = txt.find("WPQ2SchemaData") - 5
    assert 15 < start < end, "Table data markers not found in page"
    rows = json.loads(txt[start:end])["Row"]
    # Assert that the table is in descending date order (using ID as proxy for date)
    assert all(a["ID"] > b["ID"] for a, b in zip(rows, rows[1:])), "Not sorted desc"
    # Pick the first item, assuming the table is sorted in descending date order
    return f"{BASE_URL}/{rows[0]['FileLeafRef']}/{EXCEL_FILENAME}"


def crawl(context: Context) -> None:
    # First we find the link to the excel file
    excel_url = crawl_excel_url(context)
    _, _, _, path = zyte_api.fetch_resource(
        context, filename="list.xlsx", url=excel_url, geolocation="us"
    )
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)
    wb = load_workbook(path, read_only=True)
    sheet_names = wb.sheetnames
    assert wb.active is not None, "No active sheet found"
    for item in h.parse_xlsx_sheet(context, wb.active, skiprows=26):
        crawl_item(item, context)
    sheet_names.remove(wb.active.title)
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        if not (sheet.max_row == 1 and sheet.max_column == 1 and not sheet["A1"].value):
            context.log.warning(f"Sheet {sheet_name} is not empty")
