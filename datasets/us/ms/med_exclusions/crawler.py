import openpyxl
from rigour.mime.types import XLSX

from zavod import Context, Entity, helpers as h

INDEFINITE_END_DATE = "2999-12-31"


def crawl_detail(
    context: Context, entity: Entity, row: dict[str, str], npi: str
) -> None:
    entity.add("country", "us")
    entity.add("npiCode", npi.split(" "))
    entity.add("idNumber", row.pop("medicaid_id"))
    entity.add("sector", row.pop("provider_type", None))
    entity.add("sector", row.pop("speciality", None))

    address = h.make_address(
        context,
        street=row.pop("address_line_1"),
        street2=row.pop("address_line_2"),
        city=row.pop("city"),
        state=row.pop("state"),
        postal_code=row.pop("zipcode"),
        country_code="us",
    )
    h.copy_address(entity, address)

    sanction = h.make_sanction(context, entity)
    h.apply_date(sanction, "startDate", row.pop("termination_effective_date"))
    end_date = row.pop("termination_end_date")
    if end_date and end_date != INDEFINITE_END_DATE:
        h.apply_date(sanction, "endDate", end_date)
    sanction.add("reason", row.pop("termination_reason"))
    sanction.add("provisions", row.pop("sanction_type"))
    sanction.add("description", row.pop("additional_notes"))

    if h.is_active(sanction):
        entity.add("topics", "debarment")

    context.emit(entity)
    context.emit(sanction)

    context.audit_data(row, ignore=["column_0", "exclusion_period", "role"])


def crawl_item(row: dict[str, str], context: Context) -> None:
    org_name = row.pop("organization_name")
    first_name = row.pop("first_name")
    middle_name = row.pop("middle_name")
    last_name = row.pop("last_name")
    suffix = row.pop("suffix")
    dob = row.pop("date_of_birth")
    npi = row.pop("npi")
    entity_type = row.pop("entity_type")

    org_entity = None
    if org_name:
        org_entity = context.make("Organization")
        org_entity.id = context.make_id(org_name, npi)
        org_entity.add("name", org_name)
        if entity_type == "Organization":
            org_entity.add("sector", row.pop("provider_type"))
            org_entity.add("sector", row.pop("speciality"))
        crawl_detail(context, org_entity, row.copy(), npi)

    person_entity = None
    if first_name or last_name:
        person_entity = context.make("Person")
        person_entity.id = context.make_id(first_name, last_name, npi, dob)
        h.apply_name(
            person_entity,
            first_name=first_name,
            middle_name=middle_name,
            last_name=last_name,
            suffix=suffix,
        )
        person_entity.add("npiCode", npi.split(" ") if npi else [])
        h.apply_date(person_entity, "birthDate", dob)
        if entity_type == "Individual":
            person_entity.add("sector", row.pop("provider_type"))
            person_entity.add("sector", row.pop("speciality"))

        crawl_detail(context, person_entity, row.copy(), npi)

    if org_entity is not None and person_entity is not None:
        link = context.make("UnknownLink")
        link.id = context.make_id(org_entity.id, person_entity.id)
        link.add("subject", person_entity)
        link.add("object", org_entity)
        context.emit(link)


def crawl_data_url(context: Context) -> str:
    doc = context.fetch_html(context.data_url, absolute_links=True)
    url = h.xpath_string(doc, "//a[contains(text(), 'Sanctioned Provider List')]/@href")
    assert url is not None, "Could not find Excel file URL"
    return url


def crawl(context: Context) -> None:
    data_url = crawl_data_url(context)
    path = context.fetch_resource("source.xlsx", data_url)
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    workbook: openpyxl.Workbook = openpyxl.load_workbook(path, read_only=True)
    assert "Sheet1" in workbook.sheetnames, workbook.sheetnames
    header_lookup = context.dataset.lookups["headers"]
    for row in h.parse_xlsx_sheet(
        context,
        sheet=workbook["Sheet1"],
        skiprows=8,
        header_lookup=header_lookup,
    ):
        strings_row = {
            k: str(v).strip() if v is not None else "" for k, v in row.items()
        }
        crawl_item(strings_row, context)
