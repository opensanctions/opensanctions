from typing import Dict
from rigour.mime.types import XLSX
from openpyxl import load_workbook

from zavod import Context, helpers as h
from zavod.shed.zyte_api import fetch_html, fetch_resource


def crawl_item(row: Dict[str, str], context: Context):
    name = row.pop("provider_name")
    npi = row.pop("npi")
    if not name:
        return

    entity = context.make("LegalEntity")
    entity.id = context.make_id(name, npi)
    entity.add("name", name)
    entity.add("country", "us")
    entity.add("npiCode", h.multi_split(npi, ["; ", "&", " and "]))
    entity.add("alias", row.pop("doing_business_as_name"))

    sanction = h.make_sanction(context, entity)
    h.apply_date(sanction, "startDate", row.pop("termination_effective_date"))
    sanction.add("reason", row.pop("termination_authority"))
    h.apply_date(sanction, "endDate", row.pop("reinstatement_effective_date", None))
    is_active = h.is_active(sanction)
    if is_active:
        entity.add("topics", "debarment")

    context.emit(entity, target=is_active)
    context.emit(sanction)

    context.audit_data(row)


def crawl_excel_url(context: Context):
    file_xpath = ".//*[@about='/provider-termination']"
    doc = fetch_html(context, context.data_url, file_xpath)
    return doc.find(file_xpath).find(".//a").get("href")


def crawl(context: Context) -> None:
    # First we find the link to the excel file
    excel_url = crawl_excel_url(context)
    _, _, _, path = fetch_resource(
        context, "source.xlsx", excel_url, expected_media_type=XLSX
    )
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    wb = load_workbook(path, read_only=True)

    # Currently terminated providers
    for item in h.parse_xlsx_sheet(context, wb["Termination List"]):
        crawl_item(item, context)

    # Providers that where terminated but are now reinstated
    if "Reinstated Providers" in wb.sheetnames:
        for item in h.parse_xlsx_sheet(context, wb["Reinstated Providers"]):
            crawl_item(item, context)
    assert len(set(wb.sheetnames) - {"Termination List", "Reinstated Providers"}) == 0
