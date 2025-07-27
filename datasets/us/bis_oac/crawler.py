from csv import DictReader
from normality import squash_spaces, stringify
from openpyxl import load_workbook
from rigour.mime.types import XLS, CSV
from typing import Dict
from urllib.parse import urljoin
from email.message import Message

from zavod import Context
from zavod import helpers as h
from zavod.archive import dataset_data_path


def parse_header(header: str) -> str:
    m = Message()
    m["header"] = header
    return m


def crawl_row(context: Context, row: Dict[str, str]):
    requester = row.pop("REQUESTER").strip()
    # Skip the footer
    if "this list is provided to assist" in requester.lower():
        return
    requesting_country = row.pop("REQUESTING COUNTRY").strip()
    date_listed = row.pop("DATE LISTED").strip()
    if not requester and not requesting_country and not date_listed:
        return
    if not requester or not requesting_country:
        context.log.warning("Missing requester, requesting country", row=row)
        return

    entity = context.make("Company")
    entity.id = context.make_id(requester, requesting_country)
    entity.add("name", requester)
    entity.add("country", requesting_country)
    entity.add("topics", "export.risk")
    sanction = h.make_sanction(context, entity)
    h.apply_date(sanction, "listingDate", date_listed)

    context.emit(entity)
    context.emit(sanction)


def crawl_csv(context: Context, path: str, encoding: str):
    with open(path, encoding=encoding) as fh:
        # Read the file all at once to trigger any encoding errors
        # before we start emitting entities.
        lines = fh.readlines()

    # Skip the first three lines
    for row in DictReader(lines[3:]):
        crawl_row(context, row)


def crawl_xls(context: Context, path: str):
    wb = load_workbook(path, read_only=True)
    for sheet in wb.worksheets:
        headers = [
            squash_spaces(str(c.value))
            for c in list(sheet.iter_rows(min_row=1, max_row=1))[0]
        ]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = {headers[i]: stringify(row[i]) for i in range(len(headers))}
            crawl_row(context, data)


def fetch_file_url(context: Context) -> str:
    params = {"_": context.data_time.date().isoformat()}
    doc = context.fetch_html(context.data_url, params=params)
    for link in doc.xpath(".//a[text() = 'Requester List (CSV)']"):
        return urljoin(context.data_url, link.get("href"))
    raise ValueError("Could not find tabular file on the website")


def crawl(context: Context):
    context.log.info("Fetching data from the source")
    url = fetch_file_url(context)
    data_path = dataset_data_path(context.dataset.name)
    data_path.mkdir(parents=True, exist_ok=True)
    with context.http.request(method="GET", url=url, stream=True) as res:
        res.raise_for_status()
        mime_type = res.headers["content-type"].split(";")[0]
        if mime_type == XLS:
            extension = ".xls"
        elif mime_type == CSV:
            extension = ".csv"
        else:
            raise ValueError(f"Unsupported guessed file format: {mime_type}")
        path = data_path / ("source" + extension)
        with open(path, "wb") as fh:
            for chunk in res.iter_content(chunk_size=8192 * 10):
                fh.write(chunk)
    context.export_resource(path, mime_type, title=context.SOURCE_TITLE)
    if mime_type == XLS:
        crawl_xls(context, path)
    elif mime_type == CSV:
        try:
            crawl_csv(context, path, "utf-8-sig")
        except UnicodeDecodeError as e:
            context.log.info(f"Failed to decode CSV file as utf-8-sig: {e}")
            # https://en.wikipedia.org/wiki/%C3%9C as 9A
            crawl_csv(context, path, "CP437")
