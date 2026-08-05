import re
from rigour.mime.types import XLSX
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from zavod import Context, helpers as h
from zavod.extract.zyte_api import fetch_resource

REGEX_DBA = re.compile(r"\bdba\b", re.IGNORECASE)
# First cell of the header row, used to skip the notes above the table.
HEADER_FIRST_CELL = "First Name"
# The notes above the table have never been more than a handful of rows.
MAX_PREAMBLE_ROWS = 10


def count_preamble_rows(sheet: Worksheet) -> int:
    """Count the rows preceding the header row of the table.

    The source publishes a varying number of notes rows above the table, so the
    header is located by its first cell rather than assumed at a fixed offset.
    """
    for index, row in enumerate(
        sheet.iter_rows(max_row=MAX_PREAMBLE_ROWS, values_only=True)
    ):
        cell = row[0]
        if isinstance(cell, str) and cell.strip() == HEADER_FIRST_CELL:
            return index
    raise ValueError(f"Could not find header row starting with {HEADER_FIRST_CELL!r}")


def crawl_item(row: dict[str, str | None], context: Context) -> None:
    period = row.pop("timeframe_of_term_exclusion")
    if not context.lookup("period", period):
        context.log.warning("Unexpected exclusion period", period=period, row=row)

    if (first_name := row.pop("first_name")) and first_name != "N/A":
        entity = context.make("Person")
        entity.id = context.make_id(
            first_name, row.get("npi"), row.get("last_name_or_practice_name")
        )
        h.apply_name(
            entity,
            first_name=first_name,
            last_name=row.pop("last_name_or_practice_name"),
        )
    else:
        raw_business_name = row.pop("last_name_or_practice_name")
        entity = context.make("Company")
        entity.id = context.make_id(raw_business_name, row.get("npi"))

        if raw_business_name and "Owner:" in raw_business_name:
            parts = re.split(r"\s*Owner:\s*", raw_business_name)
            business_name: str | None = str(parts[0].strip())
            owner_name = parts[1].strip()
            owner = context.make("Person")
            owner.id = context.make_id(owner_name)
            owner.add("name", owner_name)
            link = context.make("Ownership")
            link.id = context.make_id(entity.id, owner.id)
            link.add("asset", entity)
            link.add("owner", owner)
            context.emit(owner)
            context.emit(link)
        else:
            business_name = raw_business_name

        assert business_name is not None
        names = REGEX_DBA.split(business_name)
        entity.add("name", names[0])
        entity.add("alias", names[1:])

    if row.get("npi") and row.get("npi") != "N/A":
        entity.add("npiCode", h.multi_split(row.pop("npi"), "; "))
    else:
        row.pop("npi")

    entity.add("country", "us")
    entity.add("topics", "debarment")
    license = row.pop("license")
    if license and license != "N/A":
        entity.add("description", "License number: " + license)

    sanction = h.make_sanction(context, entity)
    h.apply_date(sanction, "startDate", row.pop("effective_date"))
    sanction.add("reason", row.pop("reason_for_term_exclusion"))

    context.emit(entity)
    context.emit(sanction)

    context.audit_data(row)


def crawl(context: Context) -> None:
    _, _, _, path = fetch_resource(
        context, "source.xlsx", context.data_url, expected_media_type=XLSX
    )
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    wb = load_workbook(path, read_only=True)

    sheet_names = wb.sheetnames

    assert wb.active is not None
    skiprows = count_preamble_rows(wb.active)
    for item in h.parse_xlsx_sheet(context, wb.active, skiprows=skiprows):
        crawl_item(item, context)

    sheet_names.remove(wb.active.title)

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        if not (sheet.max_row == 1 and sheet.max_column == 1 and not sheet["A1"].value):
            context.log.warning(f"Sheet {sheet_name} is not empty")
