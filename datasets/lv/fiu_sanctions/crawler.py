from typing import Optional

from lxml.etree import _Element
from normality import slugify
from openpyxl import load_workbook
from rigour.mime.types import XLSX, XML
from followthemoney.types import registry

from zavod import Context, Entity
from zavod import helpers as h


FROZEN_ASSETS_URL = "https://sankcijas.fid.gov.lv/uploads/sankciju_subjekti_tabula.xlsx"


def crawl_person(context: Context, node: _Element) -> Optional[Entity]:
    entity = context.make("Person")
    entity.id = context.make_slug("person", node.findtext(".//Id"))

    name_node = node.find(".//Name")
    if name_node is None:
        context.log.error("No name for person", id=entity.id, node=node)
        return None

    h.apply_name(
        entity,
        full=name_node.findtext(".//WholeName"),
        first_name=name_node.findtext(".//FirstName"),
        middle_name=name_node.findtext(".//MiddleName"),
        last_name=name_node.findtext(".//LastName"),
    )

    gender = node.findtext(".//Gender")
    if gender is not None and gender.lower() == "v":
        entity.add("gender", "male")
    else:
        context.log.warn("Unknown gender", entity=entity, gender=gender)

    birth_node = node.find(".//Birth")
    if birth_node is not None:
        birth_date = birth_node.findtext(".//BirthDate")
        birth_country = birth_node.findtext(".//BirthCountry")
        birth_country_code = birth_node.findtext(".//BirthCountryIso2Code")

        entity.add("birthDate", birth_date)
        entity.add("birthPlace", birth_country)
        entity.add("birthCountry", birth_country_code)

    citizenship_node = node.find(".//Citizen")
    if citizenship_node is not None:
        nationality = citizenship_node.findtext(".//CitizenCountry")
        entity.add("nationality", nationality)

    for alias_node in node.findall(".//Alias"):
        h.apply_name(
            entity,
            full=alias_node.findtext(".//AliasWholeName"),
            first_name=alias_node.findtext(".//AliasFirstName"),
            middle_name=alias_node.findtext(".//AliasMiddleName"),
            last_name=alias_node.findtext(".//AliasLastName"),
            alias=True,
        )

    for document_node in node.findall(".//Document"):
        document_type = document_node.findtext(".//DocumentType")
        if document_type is None:
            context.log.error(
                "Document has no type", document=document_node, entity=entity
            )
            continue
        document_number = document_node.findtext(".//DocumentNumber")
        document_country_code = document_node.findtext(".//DocumentCountryIso2Code")

        is_passport = False
        if document_type.lower() == "pase":
            entity.add("passportNumber", document_number)
            is_passport = True
        else:
            context.log.warn(f"Unknown document ID - {document_type}")

        ident = h.make_identification(
            context,
            entity,
            number=document_number,
            country=document_country_code,
            doc_type=document_type,
            passport=is_passport,
        )
        if ident is not None:
            context.emit(ident)

    return entity


def crawl_organization(context: Context, node: _Element) -> Optional[Entity]:
    entity = context.make("Organization")
    entity.id = context.make_slug("org", node.findtext(".//Id"))

    company_name = node.find(".//Name")
    if company_name is None:
        context.log.error("Company has no name", entity=entity)
        return None
    entity.add("name", company_name.findtext(".//WholeName"))

    alias_node = node.find(".//Alias")
    if alias_node is not None:
        alias_full = alias_node.findtext(".//AliasWholeName")
        entity.add("alias", alias_full)

    address_nodes = node.findall(".//Address")
    for address_node in address_nodes:
        whole_address = address_node.findtext(".//AddressWhole")
        address_remark = address_node.findtext(".//AddressRemark")
        street = address_node.findtext(".//AddressStreet")
        city = address_node.findtext(".//AddressCity")
        country = address_node.findtext(".//AddressCountry")

        address_entity = h.make_address(
            context,
            full=whole_address,
            remarks=address_remark,
            street=street,
            city=city,
            country=country,
        )
        h.apply_address(context, entity, address_entity)
    return entity


def crawl_sanctions_xml(context: Context):
    path = context.fetch_resource("source.xml", context.data_url)
    context.export_resource(path, XML, title=context.SOURCE_TITLE)
    doc = context.parse_resource_xml(path)
    clean_doc = h.remove_namespace(doc)

    for node in clean_doc.findall(".//Entity"):
        entity_type = node.findtext(".//Type")
        if entity_type is None:
            context.log.error("No entity type", node=node)
            continue

        if entity_type.lower() == "fp":
            entity = crawl_person(context, node)
        elif entity_type.lower() == "jp":
            entity = crawl_organization(context, node)
        else:
            context.log.error("Invalid entity type", type=entity_type)
            continue

        if entity is None:
            continue

        entity.add("topics", "sanction")
        sanction = h.make_sanction(context, entity)
        sanction.add("sourceUrl", node.findtext(".//Link"))
        sanction.add("startDate", node.findtext(".//ListedOn"))
        sanction.add("program", node.findtext(".//Program"))
        sanction.add("reason", node.findtext(".//Remark"))

        context.emit(entity)
        context.emit(sanction)


def crawl_listed_names(
    context: Context,
    frozen_asset: Entity,
    frozen_assets_type: str,
    listed_names: str,
):
    for listed_name in h.multi_split(listed_names, ["\n", ";"]):
        entity = context.make("LegalEntity")
        entity.id = context.make_id(listed_name)
        entity.add("name", listed_name)

        link = context.make("UnknownLink")
        link.id = context.make_id(entity.id, frozen_assets_type, frozen_asset.id)
        link.add("subject", entity)
        link.add("object", frozen_asset)
        link.add("role", frozen_assets_type)

        context.emit(entity)
        context.emit(link)


def crawl_assets_xlsx(context: Context):
    path = context.fetch_resource("source.xlsx", FROZEN_ASSETS_URL)
    context.export_resource(path, XLSX, title="Assets of listed entities")
    wb = load_workbook(path, read_only=True)
    assert set(wb.sheetnames) == {"Dati", "Informācija par datiem"}

    for row in h.parse_xlsx_sheet(
        context, wb["Dati"], header_lookup=context.get_lookup("columns")
    ):
        name = row.pop("freeze_subject_name")
        birth_date_or_id = row.pop("birth_date_or_id")
        id_values = [name, birth_date_or_id]
        countries = row.pop("nationality").split(",")

        if birth_date_or_id and registry.date.clean_text(birth_date_or_id):
            entity = context.make("Person")
            entity.id = context.make_id(*id_values)
            entity.add("birthDate", birth_date_or_id)
            entity.add("nationality", countries)
        else:
            entity = context.make("LegalEntity")
            entity.id = context.make_id(*id_values)
            entity.add("registrationNumber", birth_date_or_id)
            entity.add("country", countries)

        entity.add("name", name)
        sanction = h.make_sanction(context, entity)
        h.apply_date(sanction, "startDate", row.pop("first_published_date"))
        frozen_assets_type = row.pop("frozen_assets_type")
        sanction.add("provisions", frozen_assets_type)
        sanction.add("reason", row.pop("reason"))
        sanction.add("program", h.multi_split(row.pop("law_references"), ["\n", ";"]))

        listed_names_str = row.pop("listed_subject_names")
        if slugify(name) == slugify(listed_names_str):
            # We're re-stating that whatever assets of this entity can be identified
            # should be consideered frozen.
            entity.add("topics", "sanction")
        else:
            # We're primarily listing identified assets of this entity
            entity.add("topics", "asset.frozen")
            crawl_listed_names(context, entity, frozen_assets_type, listed_names_str)

        context.emit(entity)
        context.emit(sanction)

        context.audit_data(row, ["start_dates", "fiu_verified"])


def crawl(context: Context):
    crawl_sanctions_xml(context)
    crawl_assets_xlsx(context)
