from dataclasses import dataclass

from openpyxl import load_workbook
from rigour.mime.types import XLSX

from zavod import Context, helpers as h
from zavod.stateful.positions import categorise, OccupancyStatus


@dataclass(frozen=True, kw_only=True)
class Resource:
    country: str
    lang: str
    filename: str
    file_pattern: str


RESOURCES = [
    Resource(
        country="dk", lang="dan", filename="dk.xlsx", file_pattern="pep_listen.xlsx"
    ),
    Resource(country="fo", lang="fao", filename="fo.xlsx", file_pattern="pep_listen-f"),
    Resource(country="gl", lang="kal", filename="gl.xlsx", file_pattern="pep_listen-g"),
]

# Categories where the "Stillingsbetegnelse" (job title) column does not hold a
# position but the member's *party*. Under these category headers every member
# row carries a party in that column rather than a role:
#   - "Medlemmer af Folketinget"        -> e.g. "Socialdemokratiet" (DK parliament)
#   - "Medlemmer af Europaparlamentet"  -> e.g. "Radikale Venstre" (EU parliament)
#   - "Medlemmer af Lagtinget ..."      -> e.g. "Fólkaflokkurin" (FO parliament; the
#                                          header also names substitutes/auditors)
#   - "Medlemmer af partiernes styrelsesorganer ..." -> e.g. "Siumut" (party boards,
#                                          present in the DK, FO and GL lists)
# For these we can't use the column as a position, so we derive a generic Danish
# position name from the category and record the party on `Person.political`
# instead. Each tuple is (substring to find in the category header, Danish
# position name). Matched by substring because the source appends statutory
# references and extra wording to some headers (e.g. the Lagting one above).
#
# The Greenlandic Inatsisartut is deliberately absent: there the column carries
# the real position ("Medlem af Inatsisartut", "Medlem af Revisionsudvalget" for
# the audit committee, ...), so those rows fall through to the default branch and
# the column is used as-is. (A substring match on "Inatsisartut" would also catch
# the "Medlemmer af Inatsisartuts Revisionsudvalg" header and mislabel auditors.)
PARTY_IN_JOB_TITLE_COLUMN = [
    ("Medlemmer af Folketinget", "Medlem af Folketinget"),
    ("Medlemmer af Europaparlamentet", "Medlem af Europaparlamentet"),
    ("Medlemmer af Lagtinget", "Medlem af Lagtinget"),
    ("partiernes styrelsesorganer", "Medlem af et partis styrelsesorgan"),
]


def get_position_and_party(category: str, job_title: str) -> tuple[str, str | None]:
    """Resolve a current-PEP row to a (Danish position name, party) tuple.

    The position name is always Danish; it is translated to English by
    `make_position(translate_name=True)`. The party is set only for the
    categories listed in `PARTY_IN_JOB_TITLE_COLUMN`, where the source puts the
    party in the job title column; it is `None` everywhere else.
    """
    for needle, position_name in PARTY_IN_JOB_TITLE_COLUMN:
        if needle in category:
            return position_name, job_title
    # Everywhere else the job title is the actual position: ministers, permanent
    # secretaries (departementschefer), special advisers, auditors, judges,
    # ambassadors, officers, state-owned-enterprise boards, Greenlandic
    # Inatsisartut members and their audit committee, etc.
    return job_title, None


def crawl_current_pep_item(
    context: Context, *, country: str, lang: str, row: dict[str, str | None]
) -> None:
    first_name = row.pop("first_name")
    last_name = row.pop("last_name")

    entity = context.make("Person")
    entity.id = context.make_slug(first_name, last_name)
    h.apply_name(entity, first_name=first_name, last_name=last_name)
    entity.add("citizenship", country)

    birth_date = row.pop("birth_date")
    h.apply_date(entity, "birthDate", birth_date.strip() if birth_date else None)
    category = row.pop("category")
    job_title = row.pop("job_title")
    assert job_title is not None, row
    assert category is not None, row
    # Most categories put the role in the job title column, but parliamentary and
    # party-board categories put the member's party there instead; resolve both.
    position_name, party = get_position_and_party(category, job_title)
    # Safety net: a party name leaking into the position column means a
    # parliamentary/party-board category header changed and stopped matching.
    assert position_name != "Socialdemokratiet", row
    if party is not None:
        entity.add("political", party, lang=lang)

    position = h.make_position(
        context, position_name, country=country, lang=lang, translate_name=True
    )
    categorisation = categorise(context, position, default_is_pep=True)

    listing_date = row.pop("listing_date")
    assert listing_date is not None, row
    occupancy = h.make_occupancy(
        context,
        entity,
        position,
        True,
        start_date=listing_date.strip(),
        categorisation=categorisation,
    )

    if occupancy:
        context.emit(entity)
        context.emit(position)
        context.emit(occupancy)

    context.audit_data(row, ignore=["new_on_list"])


def crawl_old_pep_item(
    context: Context, *, country: str, lang: str, row: dict[str, str | None]
) -> None:
    last_name = row.pop("last_name")
    first_name = row.pop("first_name")

    entity = context.make("Person")
    entity.id = context.make_slug(first_name, last_name)
    h.apply_name(entity, first_name=first_name, last_name=last_name)
    entity.add("citizenship", country)

    h.apply_date(entity, "birthDate", row.pop("birth_date"))

    job_title = row.pop("job_title")
    assert job_title is not None, row
    position = h.make_position(
        context, job_title, country=country, lang=lang, translate_name=True
    )
    # This sheet holds former PEPs. When the source fills in a removal date we
    # use it as the occupancy end date. Some entries have it left blank; those
    # people are still former PEPs by virtue of appearing in this sheet, so we
    # record the occupancy as ended without a specific end date.
    removal_date = row.pop("removal_date")
    occupation = h.make_occupancy(
        context,
        entity,
        position,
        no_end_implies_current=False,
        end_date=removal_date.strip() if removal_date is not None else None,
        status=OccupancyStatus.ENDED if removal_date is None else None,
        categorisation=categorise(context, position, default_is_pep=True),
    )

    if occupation:
        context.emit(entity)
        context.emit(position)
        context.emit(occupation)

    context.audit_data(row)


def crawl(context: Context) -> None:
    doc = context.fetch_html(context.data_url, absolute_links=True)

    for resource in RESOURCES:
        link = h.xpath_string(
            doc,
            f'//h2[strong[contains(text(), "PEP-liste ")]]/following-sibling::*//a[contains(@href, "{resource.file_pattern}")]/@href',
        )
        path = context.fetch_resource(resource.filename, link)
        context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

        wb = load_workbook(path, read_only=True)

        # Crawl old PEP list
        old_sheet = (
            "Tidligere PEP'ere"
            if "Tidligere PEP'ere" in wb.sheetnames
            else "Tidligere PEP'er"
        )
        for row in h.parse_xlsx_sheet(
            context, wb[old_sheet], header_lookup=context.get_lookup("columns")
        ):
            crawl_old_pep_item(
                context, country=resource.country, lang=resource.lang, row=row
            )

        new_sheet = (
            "Nuværende PEP'ere"
            if "Nuværende PEP'ere" in wb.sheetnames
            else "Nuværende PEP'er"
        )
        # Crawl current PEP list. The sheet groups members under category header
        # rows: a header row is blank except for the category name in the first
        # column, and applies to every member row beneath it until the next
        # header. We carry the current header in `category` and stamp it onto
        # each member row.
        category = None
        for row in h.parse_xlsx_sheet(
            context, wb[new_sheet], 1, header_lookup=context.get_lookup("columns")
        ):
            if not any([row["category"], row["first_name"], row["last_name"]]):
                continue
            # Meaning "No boards"
            if row["last_name"] == "Ingen styrelser etc.":
                continue

            # A row with only the category column populated is a header row:
            # remember it as the running category and move on.
            if row["category"] and sum(1 if v else 0 for v in list(row.values())) == 1:
                category = row["category"]
                continue
            # Otherwise it's a member row: attach the most recent header.
            else:
                row["category"] = category
            assert row["category"] is not None, row

            crawl_current_pep_item(
                context, country=resource.country, lang=resource.lang, row=row
            )

        assert len(wb.sheetnames) == 2, wb.sheetnames
