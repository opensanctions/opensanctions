import re
import xlrd  # type: ignore
import string
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.cell import Cell
from rigour.mime.types import XLSX, XLS
from urllib.parse import urljoin
from typing import Dict, List, Optional
from normality import squash_spaces, stringify
from normality.cleaning import decompose_nfkd
from followthemoney.types.identifier import IdentifierType

from zavod import Context, Entity
from zavod import helpers as h

# Match non-brackets inside an opening and closing pair of brackets
BRACKETED = re.compile(r"([(（][^\(\)]*[)）]|\[[^\[\]]*\])")

SPLITS = ["(%s)" % char for char in string.ascii_lowercase]
SPLITS = SPLITS + ["（%s）" % char for char in string.ascii_lowercase]
# WTF full-width brackets!
SPLITS = SPLITS + ["（a）", "（b）", "（c）", "\n"]
SPLITS = SPLITS + ["(i)", "(ii)", "(iii)", "(iv)", "(v)", "(vi)", "(vii)", "(viii)"]
SPLITS = SPLITS + ["; a.k.a.", "; a.k.a ", ", a.k.a.", ", f.k.a."]

ALIAS_SPLITS = SPLITS + ["; "]

DATE_SPLITS = SPLITS + [
    "、",
    "；",
    "又は",  # or
    "又は",  # or
    "または",  # or
    "生",  # living
    "に改訂",  # revised to
    "改訂",  # revised
    "日",  # date
    "及び",  # and
    "修正",  # fix
]
# Date of revision | revision | part of an OR phrase
DATE_CLEAN = re.compile(r"(\(|\)|（|）| |改訂日|改訂|まれ)")


def note_long_ids(entity, identifiers: List[str]):
    for identifier in identifiers:
        if len(identifier) > IdentifierType.max_length:
            entity.add("notes", identifier)


def str_cell(cell: Cell) -> Optional[str]:
    value = cell.value
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return str(value).lower()
    return stringify(value)


def parse_date(text: List[str]) -> List[str]:
    dates: List[str] = []
    for date_ in h.multi_split(text, DATE_SPLITS):
        parsed = h.convert_excel_date(date_)
        if parsed is not None:
            dates.append(parsed)
            continue
        cleaned = DATE_CLEAN.sub("", date_)
        if cleaned:
            normal = decompose_nfkd(cleaned)
            if normal:
                dates.append(normal)
    return dates


def parse_names(names: List[str]) -> List[str]:
    cleaned = []
    # We split on numbering e.g. (a), (b) when reading the rows of the excel file
    # so we might have split a cell-wide opening and closing parenthesis
    split_open = len([n.startswith("(") and n.count(")") == 0 for n in names])
    split_close = len([n.endswith(")") and n.count("(") == 0 for n in names])
    split_bracketed = split_open == split_close
    for name in names:
        # (a.k.a.:
        # Full width colon. Yes really.
        name = re.sub(r"[(（]a\.k\.a\.? ?[:：]? ?", "", name)
        name = re.sub(r"[（(]original script[:：]\s*(.*?)[）)]", r"\1", name)
        # It's come up in the opposite order after a U+202B right-to-left embedding char
        name = re.sub(r"[（(](.*?)\s*[:：]original script[）)]", r"\1", name)
        name = re.sub(r"[（(]f.k.a.:\s*(.*?)[）)]", r"\1", name)

        # in Excel, it has this value as (previously listed as), (Previously listed as some name)
        name = name.replace("(previously listed as)", "")
        name = name.replace("(formerly listed as", "")
        name = name.replace("a.k.a., the following three aliases:", "")
        if split_bracketed and name.count("(") == 0 and name.endswith(")"):
            name = name.rstrip(")")
        if split_bracketed and name.count(")") == 0 and name.startswith("("):
            name = name.lstrip("(")
        # e.g: Al Qaïda au Maghreb islamique (AQMI))
        name = name.replace("))", ")")
        name = name.strip(" 、")
        # U+202B Right to left embedding indicates a part of a string will be right to left
        if name and name != "\u202b":
            cleaned.append(name)
        no_brackets = BRACKETED.sub(" ", name).strip()
        if no_brackets != name and len(no_brackets):
            cleaned.append(no_brackets)
    return cleaned


def parse_notes(context: Context, entity: Entity, notes: List[str]) -> None:
    for note in notes:
        cryptos = h.extract_cryptos(note)
        for key, curr in cryptos.items():
            wallet = context.make("CryptoWallet")
            wallet.id = context.make_slug(curr, key)
            wallet.add("currency", curr)
            wallet.add("publicKey", key)
            wallet.add("topics", "sanction")
            wallet.add("holder", entity.id)
            context.emit(wallet)

        clean = h.clean_note(note)
        entity.add("notes", clean)


def fetch_excel_url(context: Context) -> str:
    params = {"_": context.data_time.date().isoformat()}
    doc = context.fetch_html(context.data_url, params=params)
    for link in doc.findall('.//div[@class="unique-block"]//a'):
        href = urljoin(context.data_url, link.get("href"))
        if href.endswith(".xlsx") or href.endswith(".xls"):
            return href
    raise ValueError("Could not find XLS file on MoF web site")


def emit_row(context: Context, sheet: str, section: str, row: Dict[str, List[str]]):
    schema = context.lookup_value("schema", section)
    if schema is None:
        context.log.warning("No schema for section", section=section, sheet=sheet)
        return
    entity = context.make(schema)
    name_english = row.pop("name_english")
    name_japanese = row.pop("name_japanese")
    passport_number = row.pop("passport_number", [])
    id_number = row.pop("id_number", [])
    identification_number = row.pop("identification_number", [])

    entity.id = context.make_id(*name_english, *name_japanese)
    if entity.id is None:
        # context.inspect((sheet, row))
        return
    entity.add("name", parse_names(name_english), lang="eng")
    entity.add("name", parse_names(name_japanese))
    entity.add("alias", parse_names(h.multi_split(row.pop("alias", []), ALIAS_SPLITS)))
    entity.add("alias", parse_names(row.pop("known_alias", [])))
    entity.add(
        "weakAlias", parse_names(h.multi_split(row.pop("weak_alias", []), ALIAS_SPLITS))
    )
    entity.add(
        "weakAlias", parse_names(h.multi_split(row.pop("nickname", []), ALIAS_SPLITS))
    )
    entity.add("previousName", parse_names(row.pop("past_alias", [])))
    entity.add("previousName", parse_names(row.pop("old_name", [])))
    entity.add_cast("Person", "position", row.pop("position", []), lang="eng")
    if entity.schema.is_a("Person"):
        birth_date = parse_date(row.pop("birth_date", []))
        if birth_date != []:
            h.apply_dates(entity, "birthDate", birth_date)
    entity.add_cast("Person", "birthPlace", row.pop("birth_place", []))

    note_long_ids(entity, passport_number)
    entity.add_cast(
        "Person",
        "passportNumber",
        h.multi_split(passport_number, SPLITS),
    )
    note_long_ids(entity, id_number)
    entity.add("idNumber", h.multi_split(id_number, SPLITS))
    note_long_ids(entity, identification_number)
    entity.add(
        "idNumber",
        h.multi_split(identification_number, SPLITS),
    )
    parse_notes(context, entity, row.pop("other_information", []))
    parse_notes(context, entity, row.pop("details", []))
    # entity.add("notes", h.clean_note(row.pop("other_information", None)))
    # entity.add("notes", h.clean_note(row.pop("details", None)))
    entity.add("phone", row.pop("phone", []))
    entity.add("phone", row.pop("fax", []))

    for address_full in row.pop("address", []):
        address = h.make_address(context, full=address_full)
        h.apply_address(context, entity, address)

    for address_full in row.pop("where", []):
        address = h.make_address(context, full=address_full)
        h.apply_address(context, entity, address)

    title = row.pop("title", [])
    if entity.schema.is_a("Person"):
        entity.add("title", title)
    else:
        entity.add("notes", title)
    entity.add("country", row.pop("citizenship", []))
    entity.add("country", row.pop("activity_area", []))

    sanction = h.make_sanction(context, entity)
    sanction.add("program", section)
    sanction.add("reason", row.pop("root_nomination", None))
    sanction.add("reason", row.pop("reason_res1483", None))
    sanction.add("authorityId", row.pop("notification_number", None))
    h.apply_dates(sanction, "startDate", parse_date(row.pop("designated_date_un", [])))
    h.apply_dates(sanction, "startDate", parse_date(row.pop("notification_date", [])))
    h.apply_dates(sanction, "listingDate", parse_date(row.pop("publication_date", [])))

    # if len(row):
    #     context.inspect(row)
    entity.add("topics", "sanction")
    context.emit(entity)
    context.emit(sanction)


def crawl_xlsx(context: Context, url: str):
    path = context.fetch_resource("source.xlsx", url)
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

    wb = load_workbook(path, read_only=True)
    for sheet in wb.worksheets:
        row0 = [str_cell(c) for c in list(sheet.iter_rows(0, 1))[0]]
        sections = [str(c) for c in row0 if c is not None]
        section = squash_spaces(" / ".join(sections))
        if section is None:
            context.log.warning("No section found", sheet=sheet.title)
            continue
        headers = None
        for cells in sheet.iter_rows(1):
            row = [str_cell(c) for c in cells]

            # after a header is found, read normal data:
            if headers is not None:
                data: Dict[str, List[str]] = {}
                for header, cell in zip(headers, row):
                    if header is None:
                        continue
                    values = []
                    if isinstance(cell, datetime):
                        cell = cell.date()
                    for value in h.multi_split(stringify(cell), SPLITS):
                        if value is None:
                            continue
                        if value == "不明":
                            continue
                        if value is not None:
                            values.append(value)
                    data[header] = values
                emit_row(context, sheet.title, section, data)

            if not len(row) or row[0] is None:
                continue
            teaser = row[0].strip()
            # the first column of the common headers:
            if "告示日付" in teaser:  # jp: Notice date
                if headers is not None:
                    context.log.error("Found double header?", row=row)
                # print("SHEET", sheet, row)
                headers = []
                for cell in row:
                    cell = squash_spaces(cell)
                    header = context.lookup_value("columns", cell)
                    if header is None:
                        context.log.warning(
                            "Unknown column title", column=cell, sheet=sheet.title
                        )
                    headers.append(header)


def crawl_xls(context: Context, url: str):
    path = context.fetch_resource("source.xls", url)
    context.export_resource(path, XLS, title=context.SOURCE_TITLE)

    xls = xlrd.open_workbook(path)
    for sheet in xls.sheets():
        headers = None
        row0 = [h.convert_excel_cell(xls, c) for c in sheet.row(0)]
        sections = [c for c in row0 if c is not None]
        section = squash_spaces(" / ".join(sections))
        if section is None:
            context.log.warning("No section found", sheet=sheet.name)
            continue
        for r in range(1, sheet.nrows):
            row = [h.convert_excel_cell(xls, c) for c in sheet.row(r)]

            # after a header is found, read normal data:
            if headers is not None:
                data: Dict[str, List[str]] = {}
                for header, cell in zip(headers, row):
                    if header is None:
                        continue
                    values = []
                    if isinstance(cell, datetime):
                        cell = cell.date()
                    for value in h.multi_split(stringify(cell), SPLITS):
                        if value is None:
                            continue
                        if value == "不明":
                            continue
                        if value is not None:
                            values.append(value)
                    data[header] = values
                emit_row(context, sheet.name, section, data)

            if not len(row) or row[0] is None:
                continue
            teaser = row[0].strip()
            # the first column of the common headers:
            if "告示日付" in teaser:  # jp: Notice date
                if headers is not None:
                    context.log.error("Found double header?", row=row)
                # print("SHEET", sheet, row)
                headers = []
                for cell in row:
                    cell = squash_spaces(cell)
                    header = context.lookup_value("columns", cell)
                    if header is None:
                        context.log.warning(
                            "Unknown column title", column=cell, sheet=sheet.name
                        )
                    headers.append(header)


def crawl(context: Context):
    url = fetch_excel_url(context)
    if url.endswith(".xlsx"):
        crawl_xlsx(context, url)
    elif url.endswith(".xls"):
        crawl_xls(context, url)
    else:
        raise ValueError("Unknown file type: %s" % url)
