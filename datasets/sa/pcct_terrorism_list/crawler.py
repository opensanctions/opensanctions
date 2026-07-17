import json
from typing import Literal, NamedTuple

from openpyxl import load_workbook
from pydantic import BaseModel, JsonValue
from rigour.mime.types import XLSX
from zavod.entity import Entity
from zavod.extract.llm import DEFAULT_MODEL, run_typed_text_prompt
from zavod.stateful.review import (
    JSONSourceValue,
    assert_all_accepted,
    review_extraction,
)

from zavod import Context
from zavod import helpers as h

PROGRAM_KEY = "SA-UNSC1373"


class IdentityDocument(BaseModel):
    type: Literal["passport", "nationalId"]
    number: str
    country: str | None = None


class IdentityDocuments(BaseModel):
    documents: list[IdentityDocument] = []


DOCUMENT_PROMPT = """
Extract every identity document from a sanctions designation. The input is a
JSON object with a document type, one or more document numbers, and an issuing
country. A single field may list several numbers separated by semicolons, or
compound entries like "A) National identification; B) Passport" that pair each
number with a type. Return one item per distinct document number.
  - type: passport, nationalId (national identification).
  - number: a single document number, exactly as written in the source.
  - country: the issuing country of this document, if stated; otherwise null.
"""


class Relation(NamedTuple):
    """How a related entity and the edge linking it to the anchor are shaped."""

    schema: str  # schema of the related entity
    link_schema: str  # schema of the edge entity
    anchor_prop: str  # edge property pointing at the anchor entity
    related_prop: str  # edge property pointing at the related entity
    role: str  # the edge's role, describing the relationship
    topics: str | None  # optional topic set on the related entity


# The terrorist organization each designation is affiliated with.
TERROR_ORG_RELATION = Relation(
    schema="Organization",
    link_schema="UnknownLink",
    anchor_prop="subject",
    related_prop="object",
    role="Linked terrorist organization",
    topics="crime.terror",
)
# The owner of a designated organization or vessel.
OWNER_RELATION = Relation(
    schema="LegalEntity",
    link_schema="Ownership",
    anchor_prop="asset",
    related_prop="owner",
    role="Owner",
    topics=None,
)


def clean_cell(value: str | None) -> str | None:
    """Strip surrounding whitespace and treat the literal 'NA' as empty."""
    if value is None:
        return None
    value = value.strip()
    if not value or value.upper() == "NA":
        return None
    return value


def emit_related(
    context: Context,
    entity: Entity,
    relation: Relation,
    name: str | None,
    notes: str | None = None,
) -> None:
    notes = clean_cell(notes)
    if clean_cell(name) is None:
        entity.add("notes", notes)
        return

    related = context.make(relation.schema)
    related.id = context.make_id(relation.role, name)
    related.add("name", name)
    related.add("topics", relation.topics)
    related.add("notes", notes)
    context.emit(related)

    link = context.make(relation.link_schema)
    link.id = context.make_id(entity.id, relation.role, related.id)
    link.add(relation.anchor_prop, entity)
    link.add(relation.related_prop, related)
    link.add("role", relation.role)
    context.emit(link)


def emit_documents(
    context: Context,
    entity: Entity,
    doc_type: str | None,
    raw_number: str | None,
    raw_country: str | None,
) -> None:
    """Emit the identity documents described by a row's document columns.

    A row carrying exactly one value in each of the three columns - a single
    number, type and issuing country - is unambiguous and emitted directly. Any
    other shape (several numbers or countries, a compound "A) ...; B) ..." type,
    or a missing value) is handed to an LLM that splits the columns into one
    document each; nothing is emitted until a reviewer accepts that extraction.
    """
    doc_type = clean_cell(doc_type)
    raw_number = clean_cell(raw_number)
    raw_country = clean_cell(raw_country)
    if raw_number is None:
        return

    numbers = h.multi_split(raw_number, [";"])
    countries = h.multi_split(raw_country, [";"])
    types = h.multi_split(doc_type, [";"])

    # Unambiguous: one number, one type, one country -> emit without review.
    if len(numbers) == 1 and len(countries) == 1 and len(types) == 1:
        is_passport = context.lookup_value("document.type", doc_type) == "passport"
        identification = h.make_identification(
            context,
            entity,
            numbers[0],
            doc_type,
            country=countries[0],
            passport=is_passport,
        )
        if identification is not None:
            context.emit(identification)
        return

    # Ambiguous: delegate parsing to an LLM and gate emission on human review.
    # The three columns are passed together, so the model assigns each document
    # its own type and country - no code-side pairing or fallback.
    source_data: JsonValue = {
        "document_type": doc_type,
        "document_number": raw_number,
        "issuing_country": raw_country,
    }
    source_value = JSONSourceValue(
        key_parts=[v for v in (raw_number, doc_type, raw_country) if v is not None],
        label="Identity documents",
        data=source_data,
        url=context.data_url,
    )
    extraction = run_typed_text_prompt(
        context,
        prompt=DOCUMENT_PROMPT,
        string=json.dumps(source_data, ensure_ascii=False),
        response_type=IdentityDocuments,
    )
    review = review_extraction(
        context,
        source_value=source_value,
        original_extraction=extraction,
        origin=DEFAULT_MODEL,
    )
    if not review.accepted:
        # Keep the raw document text in notes until the extraction is accepted,
        # so the source information isn't dropped while awaiting review.
        entity.add("notes", ", ".join(source_value.key_parts))
        return
    for document in review.extracted_data.documents:
        # Reviewed values are tagged with the review origin so the statements
        # are attributed to the model extraction rather than the crawler.
        identification = h.make_identification(
            context,
            entity,
            document.number,
            document.type,
            country=document.country,
            passport=document.type == "passport",
            origin=review.origin,
        )
        if identification is not None:
            context.emit(identification)


def crawl_row(context: Context, schema: str, row: dict[str, str | None]) -> None:
    """
    All three sheets share this path; the headers are normalized to a common
    vocabulary by the ``columns`` lookup, and properties that don't apply to a
    given schema are added with ``quiet=True`` and silently skipped.
    """
    name = row.pop("name")
    if clean_cell(name) is None:
        context.log.warning("Row without a name", row=row)
        return
    designated_date = row.pop("designated_date")
    birth_date = row.pop("birth_date", None)

    entity = context.make(schema)
    # Strict rule: only raw source values go into the ID.
    entity.id = context.make_id(name, designated_date)
    entity.add("name", name)
    if entity.schema.is_a("Person"):
        h.apply_dates(entity, "birthDate", h.multi_split(birth_date, [";"]))
    entity.add("alias", h.multi_split(row.pop("alias", None), [";"]), quiet=True)
    entity.add("gender", row.pop("gender", None), quiet=True)
    entity.add(
        "birthPlace", h.multi_split(row.pop("birth_place", None), [";"]), quiet=True
    )
    entity.add(
        "nationality", h.multi_split(row.pop("nationality", None), [";"]), quiet=True
    )
    entity.add(
        "country", h.multi_split(row.pop("birth_country", None), [";"]), quiet=True
    )
    entity.add("flag", row.pop("flag", None), quiet=True)
    entity.add("imoNumber", row.pop("imo_number", None), quiet=True)
    entity.add("notes", clean_cell(row.pop("notes", None)))
    entity.add(
        "registrationNumber",
        h.multi_split(row.pop("reg_number", None), [";"]),
        quiet=True,
    )
    for full in h.multi_split(row.pop("address", None), [";"]):
        entity.add("address", full)
    # Emit identity documents
    emit_documents(
        context,
        entity,
        row.pop("doc_type", None),
        row.pop("doc_number", None),
        row.pop("issuing_country", None),
    )
    # Emit linked organizations
    for org in h.multi_split(row.pop("linked_org", None), [";"]):
        emit_related(context, entity, TERROR_ORG_RELATION, name=org)
    # Emit owners of designated vessels
    emit_related(
        context,
        entity,
        OWNER_RELATION,
        name=row.pop("owner_name", None),
        notes=row.pop("owner_info", None),
    )

    entity.add("topics", "crime.terror")
    entity.add("topics", "sanction")

    sanction = h.make_sanction(
        context,
        entity,
        program_key=PROGRAM_KEY,
        start_date=designated_date,
    )

    context.emit(entity)
    context.emit(sanction)

    context.audit_data(row, ignore=["no"])


def crawl(context: Context) -> None:
    path = context.fetch_resource("source.xlsx", context.data_url)
    context.export_resource(path, XLSX, title=context.SOURCE_TITLE)
    workbook = load_workbook(path, read_only=True)

    columns = context.get_lookup("columns")
    for sheet in workbook.sheetnames:
        schema = context.lookup_value("schema", sheet)
        if schema is None:
            context.log.warning("Unmapped sheet", sheet=sheet)
            continue
        for row in h.parse_xlsx_sheet(context, workbook[sheet], header_lookup=columns):
            crawl_row(context, schema, row)

    # Ambiguous document extractions await human review; warn rather than block
    # the run, since the documents are a secondary enrichment here.
    assert_all_accepted(context, raise_on_unaccepted=False)
